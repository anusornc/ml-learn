#!/usr/bin/env python3
"""Generate the 1-epoch Perceptron + MLP hand-calculation workbooks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUT_DIR = Path(__file__).resolve().parent
BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial")
BOLD = Font(name="Arial", bold=True, size=14)
HEAD = Font(name="Arial", bold=True, color="FFFFFF")
NOTE = Font(name="Arial", italic=True, color="334155", size=10)
YELLOW = PatternFill("solid", fgColor="FFF3BF")
NAVY = PatternFill("solid", fgColor="1E3A5F")
TEAL = PatternFill("solid", fgColor="0F766E")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
WRAP = Alignment(wrap_text=True, vertical="center")


def _width(ws: Worksheet, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _label(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = Font(name="Arial", bold=True)


def _input(ws: Worksheet, cell: str, value: float | int) -> None:
    ws[cell] = value
    ws[cell].font = BLUE
    ws[cell].fill = YELLOW
    ws[cell].border = THIN


def _blank(ws: Worksheet, cell: str, solved: bool, formula: str) -> None:
    ws[cell] = formula if solved else None
    ws[cell].font = BLACK
    ws[cell].fill = YELLOW
    ws[cell].border = THIN
    ws[cell].number_format = "0.0000"


def _headers(ws: Worksheet, row: int, values: list[str], fill: PatternFill) -> None:
    for idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=idx, value=value)
        cell.font = HEAD
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.border = THIN


def build_guide(ws: Worksheet) -> None:
    ws.title = "00_Guide"
    ws["A1"] = "ใบงาน 1 epoch — Perceptron (AND) และ MLP (XOR)"
    ws["A1"].font = BOLD
    ws.merge_cells("A1:F1")
    lines = [
        "ลำดับทำ: อ่าน 09-ML-NN.html → เปิด index.html (sandbox) → รัน perceptron.py / mlp_xor.py → ทำใบงานนี้",
        "เซลล์พื้นเหลือง + ตัวเลขน้ำเงิน = ค่าที่กำหนดให้ ห้ามแก้ตอนสอบ ยกเว้นอาจารย์จะเปลี่ยนโจทย์",
        "เซลล์พื้นเหลืองว่าง = ให้เติมสูตรเอง อย่าพิมพ์คำตอบลงไป",
        "Step(z) = 1 เมื่อ z ≥ 0 ไม่ใช่ z > 0 — จุด z = 0 สำคัญในโจทย์ AND แถวแรก",
        "Backprop: คำนวณ delta ทุกชั้นจากน้ำหนักเดิม แล้วค่อยอัปเดต (อย่าใช้ w ใหม่ตอนย้อนกลับ)",
        "Loss ของ MLP: L = 0.5 * (ŷ − y)² แล้วเดินลงเกรเดียนต์ w ← w − η · ∂L/∂w",
        "ไฟล์เฉลยมีสูตรครบ ใช้ตรวจหลังทำโจทย์ ไม่ใช่ตอนสอบ",
    ]
    for idx, line in enumerate(lines, start=3):
        ws[f"A{idx}"] = line
        ws[f"A{idx}"].font = Font(name="Arial")
        ws.merge_cells(f"A{idx}:F{idx}")
        ws[f"A{idx}"].alignment = WRAP
        ws.row_dimensions[idx].height = 22
    _width(ws, {"A": 110, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})


def build_perceptron(ws: Worksheet, *, solved: bool) -> None:
    ws.title = "01_Perceptron_AND"
    ws["A1"] = "AND หนึ่ง epoch ด้วย Perceptron (เริ่ม w=0, b=0, η=0.1)"
    ws["A1"].font = BOLD
    ws.merge_cells("A1:J1")
    _label(ws, "A3", "η")
    _input(ws, "B3", 0.1)
    _label(ws, "A4", "w1 เริ่ม")
    _input(ws, "B4", 0)
    _label(ws, "A5", "w2 เริ่ม")
    _input(ws, "B5", 0)
    _label(ws, "A6", "b เริ่ม")
    _input(ws, "B6", 0)
    ws["D3"] = (
        "ลำดับตัวอย่างคงที่: (0,0)→0, (0,1)→0, (1,0)→0, (1,1)→1"
        " — แถวแรก z=0 จึง ŷ=1 และ error=−1 ทำให้ bias ติดลบ"
    )
    ws["D3"].font = NOTE
    ws.merge_cells("D3:J6")
    ws["D3"].alignment = WRAP

    _headers(
        ws,
        8,
        ["ลำดับ", "x1", "x2", "y", "z", "ŷ = Step(z≥0)", "error = y−ŷ", "w1 ใหม่", "w2 ใหม่", "b ใหม่"],
        NAVY,
    )
    samples = [(1, 0, 0, 0), (2, 0, 1, 0), (3, 1, 0, 0), (4, 1, 1, 1)]
    for row, (order, x1, x2, y) in enumerate(samples, start=9):
        ws.cell(row=row, column=1, value=order).font = BLACK
        ws.cell(row=row, column=2, value=x1).font = BLUE
        ws.cell(row=row, column=3, value=x2).font = BLUE
        ws.cell(row=row, column=4, value=y).font = BLUE
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN
        if row == 9:
            w1, w2, b = "$B$4", "$B$5", "$B$6"
        else:
            prev = row - 1
            w1, w2, b = f"H{prev}", f"I{prev}", f"J{prev}"
        _blank(ws, f"E{row}", solved, f"={b}+{w1}*B{row}+{w2}*C{row}")
        _blank(ws, f"F{row}", solved, f"=IF(E{row}>=0,1,0)")
        _blank(ws, f"G{row}", solved, f"=D{row}-F{row}")
        _blank(ws, f"H{row}", solved, f"={w1}+$B$3*G{row}*B{row}")
        _blank(ws, f"I{row}", solved, f"={w2}+$B$3*G{row}*C{row}")
        _blank(ws, f"J{row}", solved, f"={b}+$B$3*G{row}")

    ws["A14"] = "ตรวจ: หลังแถว 4 ต้องได้ w1=0.1, w2=0.1, b=0.0"
    ws["A14"].font = NOTE
    ws.merge_cells("A14:J14")
    _width(ws, {get_column_letter(i): 16 for i in range(1, 11)})
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16


def build_mlp(ws: Worksheet, *, solved: bool) -> None:
    ws.title = "02_MLP_XOR"
    ws["A1"] = "XOR หนึ่งตัวอย่าง [x1=1, x2=0, y=1] โครงข่าย 2-2-1, sigmoid, L=½(ŷ−y)², η=2.0"
    ws["A1"].font = BOLD
    ws.merge_cells("A1:F1")

    specs = [
        ("A3", "x1", "B3", 1),
        ("C3", "x2", "D3", 0),
        ("E3", "y", "F3", 1),
        ("A4", "η", "B4", 2),
        ("A6", "W_h1,1", "B6", 0.5),
        ("C6", "W_h1,2", "D6", -0.5),
        ("E6", "b_h1", "F6", 0),
        ("A7", "W_h2,1", "B7", -0.5),
        ("C7", "W_h2,2", "D7", 0.5),
        ("E7", "b_h2", "F7", 0),
        ("A8", "W_o,1", "B8", 0.5),
        ("C8", "W_o,2", "D8", 0.5),
        ("E8", "b_o", "F8", 0),
    ]
    for label_cell, label, value_cell, value in specs:
        _label(ws, label_cell, label)
        _input(ws, value_cell, value)

    ws["A10"] = "Forward — เติม z แล้วใช้ =1/(1+EXP(-z)) สำหรับ sigmoid"
    ws["A10"].font = Font(name="Arial", bold=True)
    _headers(ws, 11, ["ชั้น", "สูตร z", "z", "activation", "L = 0.5*(ŷ−y)^2", ""], TEAL)
    ws["A12"] = "h1"
    ws["B12"] = "b_h1 + W_h1,1·x1 + W_h1,2·x2"
    ws["A13"] = "h2"
    ws["B13"] = "b_h2 + W_h2,1·x1 + W_h2,2·x2"
    ws["A14"] = "out"
    ws["B14"] = "b_o + W_o,1·a_h1 + W_o,2·a_h2"
    for row in (12, 13, 14):
        ws[f"A{row}"].border = THIN
        ws[f"B{row}"].border = THIN
        ws[f"B{row}"].font = Font(name="Arial", size=9)

    _blank(ws, "C12", solved, "=F6+B6*$B$3+D6*$D$3")
    _blank(ws, "D12", solved, "=1/(1+EXP(-C12))")
    _blank(ws, "C13", solved, "=F7+B7*$B$3+D7*$D$3")
    _blank(ws, "D13", solved, "=1/(1+EXP(-C13))")
    _blank(ws, "C14", solved, "=F8+B8*D12+D8*D13")
    _blank(ws, "D14", solved, "=1/(1+EXP(-C14))")
    _blank(ws, "E14", solved, "=0.5*(D14-F3)^2")

    ws["A16"] = "Backward — ใช้ weight เดิมเท่านั้น แล้วย่อยไปอัปเดตด้านล่าง"
    ws["A16"].font = Font(name="Arial", bold=True)
    _headers(ws, 17, ["delta", "สูตร", "ค่า", "", "", ""], TEAL)
    ws["A18"] = "δ_out"
    ws["B18"] = "(ŷ−y) · ŷ · (1−ŷ)"
    ws["A19"] = "δ_h1"
    ws["B19"] = "(W_o,1 · δ_out) · a_h1 · (1−a_h1)"
    ws["A20"] = "δ_h2"
    ws["B20"] = "(W_o,2 · δ_out) · a_h2 · (1−a_h2)"
    _blank(ws, "C18", solved, "=(D14-F3)*D14*(1-D14)")
    _blank(ws, "C19", solved, "=(B8*C18)*D12*(1-D12)")
    _blank(ws, "C20", solved, "=(D8*C18)*D13*(1-D13)")
    for row in (18, 19, 20):
        ws[f"A{row}"].border = THIN
        ws[f"B{row}"].border = THIN
        ws[f"B{row}"].font = Font(name="Arial", size=9)

    ws["A22"] = "อัปเดต: พารามิเตอร์ใหม่ = ของเดิม − η · δ · activation ขาเข้า"
    ws["A22"].font = Font(name="Arial", bold=True)
    _headers(ws, 23, ["พารามิเตอร์", "สูตรอัปเดต", "ค่าใหม่", "", "", ""], TEAL)
    updates = [
        ("W_o,1", "=B8-$B$4*C18*D12"),
        ("W_o,2", "=D8-$B$4*C18*D13"),
        ("b_o", "=F8-$B$4*C18"),
        ("W_h1,1", "=B6-$B$4*C19*$B$3"),
        ("W_h1,2", "=D6-$B$4*C19*$D$3"),
        ("b_h1", "=F6-$B$4*C19"),
        ("W_h2,1", "=B7-$B$4*C20*$B$3"),
        ("W_h2,2", "=D7-$B$4*C20*$D$3"),
        ("b_h2", "=F7-$B$4*C20"),
    ]
    for idx, (name, formula) in enumerate(updates):
        row = 24 + idx
        ws[f"A{row}"] = name
        ws[f"A{row}"].border = THIN
        ws[f"B{row}"] = "ของเดิม − η·δ·ขาเข้า"
        ws[f"B{row}"].font = Font(name="Arial", size=9)
        ws[f"B{row}"].border = THIN
        _blank(ws, f"C{row}", solved, formula)

    ws["A34"] = "ตรวจคร่าว ๆ: z_h1 ต้องเป็น 0.5, z_h2 ต้องเป็น −0.5, z_out ต้องเป็น 0.5, ŷ ≈ 0.6225"
    ws["A34"].font = NOTE
    ws.merge_cells("A34:F34")
    _width(ws, {"A": 16, "B": 42, "C": 16, "D": 16, "E": 22, "F": 12})


def build(solved: bool) -> Workbook:
    wb = Workbook()
    build_guide(wb.active)
    build_perceptron(wb.create_sheet(), solved=solved)
    build_mlp(wb.create_sheet(), solved=solved)
    return wb


def main() -> None:
    student = OUT_DIR / "Perceptron_MLP_1_Epoch_TH.xlsx"
    solved = OUT_DIR / "Perceptron_MLP_1_Epoch_TH_Solved.xlsx"
    build(False).save(student)
    build(True).save(solved)
    print(f"wrote {student.name} and {solved.name}")


if __name__ == "__main__":
    main()
