import openpyxl

def main():
    # Load practice file
    wb = openpyxl.load_workbook('Logistic_Regression_Practice_20_Steps_TH.xlsx', data_only=False)

    # 1. Sheet 00_Guide
    ws_guide = wb['00_Guide']
    for r in range(1, ws_guide.max_row+1):
        for c in range(1, ws_guide.max_column+1):
            v = ws_guide.cell(r, c).value
            if isinstance(v, str) and '20 รอบ' in v:
                ws_guide.cell(r, c).value = v.replace('20 รอบ', '10 รอบ')

    # 2. Sheet 01_Data
    ws_data = wb['01_Data']
    train_new = [
        (1.0, 50, 0),
        (2.0, 55, 0),
        (1.5, 65, 0),
        (2.5, 60, 0),
        (3.0, 70, 0),
        (3.5, 65, 0),
        (4.0, 72, 0),
        (4.0, 80, 1),
        (4.5, 75, 0),
        (5.0, 78, 1),
        (5.5, 82, 1),
        (6.0, 80, 1),
        (6.5, 85, 1),
        (7.0, 88, 1),
        (7.5, 92, 1),
        (8.0, 90, 1),
        (9.0, 96, 1),
        (3.5, 88, 1),
        (6.5, 62, 0),
        (8.5, 70, 1)
    ]

    test_new = [
        (1.5, 52, 0),
        (3.0, 78, 0),
        (4.5, 85, 1),
        (5.0, 60, 0),
        (5.5, 92, 1),
        (6.0, 70, 1),
        (7.0, 58, 0),
        (3.0, 92, 1),
        (8.0, 82, 1),
        (4.0, 68, 0)
    ]

    for idx, (hrs, att, y) in enumerate(train_new, start=6):
        ws_data.cell(row=idx, column=2).value = hrs
        ws_data.cell(row=idx, column=3).value = att
        ws_data.cell(row=idx, column=4).value = y

    for idx, (hrs, att, y) in enumerate(test_new, start=26):
        ws_data.cell(row=idx, column=2).value = hrs
        ws_data.cell(row=idx, column=3).value = att
        ws_data.cell(row=idx, column=4).value = y

    ws_data['K6'] = 5.0
    ws_data['K7'] = 2.5
    ws_data['K8'] = 75.0
    ws_data['K9'] = 15.0

    # 3. Sheet 02_Training_Detail
    ws_train = wb['02_Training_Detail']
    ws_train['B5'] = 10

    for r in range(1, 10):
        for c in range(1, 10):
            v = ws_train.cell(r, c).value
            if isinstance(v, str) and '20 รอบ' in v:
                ws_train.cell(r, c).value = v.replace('20 รอบ', '10 รอบ')

    ws_train.delete_rows(271, 530 - 271)

    # 4. Sheet 03_Model_Path
    ws_path = wb['03_Model_Path']
    for r in range(1, 6):
        for c in range(1, 10):
            v = ws_path.cell(r, c).value
            if isinstance(v, str) and '20 รอบ' in v:
                ws_path.cell(r, c).value = v.replace('20 รอบ', '10 รอบ')

    ws_path.delete_rows(16, 10)

    ws_path['B19'] = '=I15'
    ws_path['B20'] = '=J15'
    ws_path['B21'] = '=K15'

    for chart in ws_path._charts:
        for s in chart.series:
            if s.val and hasattr(s.val, 'numRef') and s.val.numRef:
                f = s.val.numRef.f
                if '$6:$25' in f:
                    s.val.numRef.f = f.replace('$6:$25', '$6:$15')
            if s.cat and hasattr(s.cat, 'strRef') and s.cat.strRef:
                f = s.cat.strRef.f
                if '$6:$25' in f:
                    s.cat.strRef.f = f.replace('$6:$25', '$6:$15')

    # 5. Sheet 04_Test_Prediction
    ws_test = wb['04_Test_Prediction']
    ws_test['B4'] = 0.5
    ws_test['E4'] = '=IF(03_Model_Path!B19="","",03_Model_Path!B19)'
    ws_test['H4'] = '=IF(03_Model_Path!B20="","",03_Model_Path!B20)'
    ws_test['K4'] = '=IF(03_Model_Path!B21="","",03_Model_Path!B21)'

    # 6. Sheet 06_Key_Concepts
    ws_kc = wb['06_Key_Concepts']
    for r in range(1, ws_kc.max_row+1):
        for c in range(1, ws_kc.max_column+1):
            v = ws_kc.cell(r, c).value
            if isinstance(v, str) and '20 รอบ' in v:
                ws_kc.cell(r, c).value = v.replace('20 รอบ', '10 รอบ')

    # 7. Sheet 07_Exercises - Update Questions to Set 1 (GD & Convergence Focus)
    ws_ex = wb['07_Exercises']
    new_questions = [
        "1. ในการฝึกโมเดล 10 รอบ ค่า Average Log Loss มีแนวโน้มลดลงอย่างไรในแต่ละรอบ และเริ่มชะลอตัวลงที่รอบใด?",
        "2. เมื่อสังเกตขนาดของ Gradient (grad b, grad w1, grad w2) จากรอบที่ 1 ถึงรอบที่ 10 มีแนวโน้มอย่างไร เหตุใดขนาดจึงลดลง?",
        "3. หากทดลองปรับ Learning Rate (α) ใน 02_Training_Detail เป็น 0.10 และ 0.80 การลู่เข้าของ Loss ใน 10 รอบเปลี่ยนแปลงอย่างไร?",
        "4. จากตาราง Threshold (0.3–0.7) การเปลี่ยน Threshold มีผลต่อ Precision และ Recall อย่างไร และ Threshold ใดให้ F1-score สูงสุด?",
        "5. เหตุใดการทำ Center & Scale จึงช่วยให้ Gradient Descent ใน 10 รอบนี้ลู่เข้าได้อย่างมีเสถียรภาพมากกว่าการใช้ข้อมูลดิบ?"
    ]

    for idx, q_text in enumerate(new_questions, start=13):
        ws_ex[f'A{idx}'] = q_text

    output_file = 'Logistic_Regression_Practice_10_Steps_TH.xlsx'
    wb.save(output_file)
    print(f'Saved practice file with updated questions: {output_file}')

    # Create Solved File
    for iter_num in range(1, 11):
        start_r = 11 + (iter_num - 1) * 26
        param_r = start_r + 1  # Row 12, 38, 64, ...
        data_start = start_r + 3  # Row 14, 40, 66, ...
        data_end = data_start + 19 # Row 33, 59, 85, ...
        sum_r = data_end + 1  # Row 34, 60, 86, ...
        upd_r = sum_r + 1     # Row 35, 61, 87, ...

        for r in range(data_start, data_end + 1):
            ws_train[f'E{r}'] = f'=$D${param_r} + $F${param_r}*B{r} + $H${param_r}*C{r}'
            ws_train[f'F{r}'] = f'=1/(1+EXP(-E{r}))'
            ws_train[f'G{r}'] = f'=F{r}-D{r}'
            ws_train[f'H{r}'] = f'=G{r}'
            ws_train[f'I{r}'] = f'=G{r}*B{r}'
            ws_train[f'J{r}'] = f'=G{r}*C{r}'
            ws_train[f'K{r}'] = f'=-D{r}*LN(F{r})-(1-D{r})*LN(1-F{r})'

        ws_train[f'H{sum_r}'] = f'=AVERAGE(H{data_start}:H{data_end})'
        ws_train[f'I{sum_r}'] = f'=AVERAGE(I{data_start}:I{data_end})'
        ws_train[f'J{sum_r}'] = f'=AVERAGE(J{data_start}:J{data_end})'
        ws_train[f'K{sum_r}'] = f'=AVERAGE(K{data_start}:K{data_end})'

        ws_train[f'D{upd_r}'] = f'=$D${param_r} - $B$4*H{sum_r}'
        ws_train[f'F{upd_r}'] = f'=$F${param_r} - $B$4*I{sum_r}'
        ws_train[f'H{upd_r}'] = f'=$H${param_r} - $B$4*J{sum_r}'

    solved_file = 'Logistic_Regression_Practice_10_Steps_TH_Solved.xlsx'
    wb.save(solved_file)
    print(f'Saved solved file: {solved_file}')

if __name__ == '__main__':
    main()
