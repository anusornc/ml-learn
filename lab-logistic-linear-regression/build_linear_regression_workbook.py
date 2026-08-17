from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from build_logistic_regression_workbook import (
    BLACK,
    BLUE,
    GRAY,
    GREEN,
    LIGHT_GRAY,
    NAVY,
    ORANGE,
    PALE_BLUE,
    PALE_GREEN,
    PALE_ORANGE,
    PALE_PURPLE,
    PALE_RED,
    PURPLE,
    RED,
    SKY,
    TEAL,
    WHITE,
    configure_sheet,
    fill,
    font,
    grid_border,
    header_row,
    medium_navy,
    note_box,
    section,
    thin_gray,
    title,
)


OUT = Path(__file__).with_name("Linear_Regression_Step_by_Step_TH.xlsx")


def merged_block(ws, cell_range, text, color):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = text
    c.font = font(9, True, NAVY)
    c.fill = fill(color)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = Border(left=medium_navy, right=thin_gray, top=thin_gray, bottom=thin_gray)


def build_guide(wb):
    ws = wb.active
    ws.title = "00_Guide"
    configure_sheet(ws)
    title(
        ws,
        "Linear Regression — คำนวณทีละขั้นใน Excel",
        "จากข้อมูลดิบ → Gradient Descent → Prediction → Residual → MAE / MSE / RMSE / R²",
        8,
    )
    ws.freeze_panes = "A4"
    for col, width in {"A": 4, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24, "G": 24, "H": 4}.items():
        ws.column_dimensions[col].width = width

    section(ws, 4, "เป้าหมายการเรียนรู้", 2, 7)
    goals = [
        "1. อธิบายสมการ ŷ = w1x1 + w2x2 + b และความหมายของ weights กับ bias ได้",
        "2. คำนวณ Prediction, Residual, Squared Error และ Loss ทีละแถวได้",
        "3. เห็นว่า Gradient Descent ปรับ w1, w2 และ b อย่างไรในแต่ละรอบ",
        "4. คำนวณและตีความ MAE, MSE, RMSE และ R² บนข้อมูล Test ได้",
    ]
    for r, text_value in enumerate(goals, 5):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(r, 2, text_value)
        c.fill = fill(SKY if r % 2 else LIGHT_GRAY)
        c.font = font(10, False, NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = grid_border
        ws.row_dimensions[r].height = 28

    section(ws, 10, "เส้นทางการเรียนใน Workbook", 2, 7)
    steps = [
        ("01_Data", "ข้อมูล 30 คน การแบ่ง Train/Test และการปรับสเกล x1, x2"),
        ("02_Training_Detail", "คำนวณ ŷ, residual, gradient, squared error และอัปเดต w1, w2, b ทีละรอบ"),
        ("03_Model_Path", "ดูเส้นทางของค่าน้ำหนักและกราฟ Loss ตลอด 15 รอบ"),
        ("04_Test_Prediction", "ทำนายคะแนนชุด Test และวิเคราะห์ residual รายคน"),
        ("05_Evaluation", "คำนวณ MAE, MSE, RMSE, R² และ Mean Residual"),
        ("06_Key_Concepts", "อ่านคำอธิบายแนวคิดสำคัญและการตีความ coefficient"),
        ("07_Exercises", "ทดลองทำนายนักศึกษาใหม่และเปลี่ยน Learning rate"),
    ]
    for r, (name, desc) in enumerate(steps, 11):
        ws.cell(r, 2, name)
        ws.cell(r, 2).font = font(10, True, TEAL)
        ws.cell(r, 2).fill = fill(PALE_BLUE)
        ws.cell(r, 2).border = grid_border
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.cell(r, 3, desc)
        ws.cell(r, 3).font = font()
        ws.cell(r, 3).border = grid_border
        ws.cell(r, 3).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 26

    section(ws, 19, "สูตรหลัก", 2, 7)
    formulas = [
        ("Prediction", "ŷ = w1x1 + w2x2 + b"),
        ("Residual", "error = ŷ - y"),
        ("Loss ต่อแถว", "0.5 × error²"),
        ("Gradient", "grad b = AVERAGE(error); grad w1 = AVERAGE(error×x1); grad w2 = AVERAGE(error×x2)"),
        ("Update", "พารามิเตอร์ใหม่ = พารามิเตอร์เดิม - learning rate × gradient"),
        ("R²", "1 - SSE / SST"),
    ]
    for r, (label, formula_text) in enumerate(formulas, 20):
        ws.cell(r, 2, label)
        ws.cell(r, 2).font = font(10, True, NAVY)
        ws.cell(r, 2).fill = fill(LIGHT_GRAY)
        ws.cell(r, 2).border = grid_border
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.cell(r, 3, formula_text)
        ws.cell(r, 3).font = font(10, False, PURPLE)
        ws.cell(r, 3).fill = fill(PALE_PURPLE)
        ws.cell(r, 3).border = grid_border
        ws.row_dimensions[r].height = 25

    note_box(
        ws,
        28,
        2,
        7,
        "จุดต่างสำคัญจาก Logistic Regression: Linear Regression ทำนายค่าต่อเนื่องโดยตรง จึงไม่มี Sigmoid, Threshold หรือ Confusion Matrix และประเมินด้วยขนาดของ Residual แทน",
        PALE_ORANGE,
    )
    ws.print_area = "A1:H30"


def build_data(wb):
    ws = wb.create_sheet("01_Data")
    configure_sheet(ws)
    title(
        ws,
        "01 • ข้อมูลและการเตรียมฟีเจอร์",
        "ทำนายคะแนนสอบจากชั่วโมงอ่านและอัตราเข้าเรียน • Train 20 คน / Test 10 คน",
        17,
    )
    ws.page_setup.orientation = "landscape"
    ws.freeze_panes = "A6"
    headers = [
        "Student ID",
        "ชั่วโมงอ่าน/วัน",
        "เข้าเรียน (%)",
        "คะแนนจริง y",
        "Split",
        "x1\n(Centered & scaled hours)",
        "x2\n(Centered & scaled attendance)",
        "คำอธิบาย",
    ]
    for c, value in enumerate(headers, 1):
        ws.cell(5, c, value)
    header_row(ws, 5, 1, 8)
    widths = {"A": 13, "B": 15, "C": 15, "D": 14, "E": 10, "F": 20, "G": 22, "H": 20, "I": 3, "J": 24, "K": 15, "L": 3, "M": 21, "N": 16, "O": 16, "P": 22, "Q": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    data = [
        ("S01", 1.0, 55, 43, "Train"), ("S02", 1.5, 60, 47, "Train"),
        ("S03", 2.0, 58, 51, "Train"), ("S04", 2.5, 65, 55, "Train"),
        ("S05", 3.0, 62, 56, "Train"), ("S06", 3.5, 70, 63, "Train"),
        ("S07", 4.0, 68, 61, "Train"), ("S08", 4.5, 72, 68, "Train"),
        ("S09", 5.0, 75, 72, "Train"), ("S10", 5.5, 78, 76, "Train"),
        ("S11", 6.0, 80, 79, "Train"), ("S12", 6.5, 82, 82, "Train"),
        ("S13", 7.0, 85, 87, "Train"), ("S14", 7.5, 88, 89, "Train"),
        ("S15", 8.0, 90, 92, "Train"), ("S16", 8.5, 92, 96, "Train"),
        ("S17", 9.0, 95, 99, "Train"), ("S18", 5.0, 65, 65, "Train"),
        ("S19", 3.0, 85, 70, "Train"), ("S20", 7.0, 65, 78, "Train"),
        ("T01", 2.0, 55, 50, "Test"), ("T02", 3.0, 75, 60, "Test"),
        ("T03", 4.0, 82, 72, "Test"), ("T04", 5.0, 65, 67, "Test"),
        ("T05", 5.0, 90, 80, "Test"), ("T06", 6.0, 72, 73, "Test"),
        ("T07", 7.0, 60, 71, "Test"), ("T08", 2.5, 95, 72, "Test"),
        ("T09", 8.0, 80, 88, "Test"), ("T10", 4.5, 70, 65, "Test"),
    ]
    for r, row in enumerate(data, 6):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
        ws.cell(r, 6, f"=(B{r}-$K$6)/$K$7")
        ws.cell(r, 7, f"=(C{r}-$K$8)/$K$9")
        ws.cell(r, 8, f'=IF(E{r}="Train","ใช้ฝึกโมเดล","เก็บไว้ประเมิน")')
        for c in range(1, 9):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = font(10, False, BLUE if c <= 5 else BLACK)
        ws.cell(r, 3).number_format = '0"%"'
        ws.cell(r, 6).number_format = "0.000"
        ws.cell(r, 7).number_format = "0.000"
        ws.cell(r, 5).fill = fill(PALE_BLUE if row[4] == "Train" else PALE_ORANGE)
        ws.row_dimensions[r].height = 23

    section(ws, 4, "ข้อมูลหลัก", 1, 8)
    section(ws, 4, "ค่าที่ใช้ปรับสเกล", 10, 11, ORANGE)
    assumptions = [("Hours center", 5.0), ("Hours scale", 2.5), ("Attendance center", 75.0), ("Attendance scale", 15.0)]
    for r, (label, value) in enumerate(assumptions, 6):
        ws.cell(r, 10, label)
        ws.cell(r, 11, value)
        for c in (10, 11):
            ws.cell(r, c).border = grid_border
            ws.cell(r, c).alignment = Alignment(vertical="center")
        ws.cell(r, 10).fill = fill(LIGHT_GRAY)
        ws.cell(r, 10).font = font(9, True, NAVY)
        ws.cell(r, 11).fill = fill(PALE_ORANGE)
        ws.cell(r, 11).font = font(10, True, BLUE)
    ws["J11"] = "สูตร"
    ws["K11"] = "x = (ค่าดิบ - center) / scale"
    for cell in (ws["J11"], ws["K11"]):
        cell.border = grid_border
        cell.alignment = Alignment(wrap_text=True)
    ws["J11"].font = font(9, True, NAVY)
    ws["J11"].fill = fill(LIGHT_GRAY)
    ws["K11"].font = font(9, False, PURPLE)
    ws["K11"].fill = fill(PALE_PURPLE)

    ws["K6"].comment = Comment("ค่าเฉลี่ย Hours ของ Train = 5.00 จึงใช้เป็น center", "OpenAI")
    ws["K7"].comment = Comment("Standard deviation จริง ≈ 2.35 ปัดเป็น 2.5 เพื่อคำนวณมือได้ง่าย", "OpenAI")
    ws["K8"].comment = Comment("ค่าเฉลี่ย Attendance ของ Train = 74.5% ปัดเป็น 75%", "OpenAI")
    ws["K9"].comment = Comment("เลือก 15 จุดเปอร์เซ็นต์เพื่อให้ 60/75/90% แปลงเป็น -1/0/+1", "OpenAI")

    section(ws, 4, "เหตุผลของการเตรียมข้อมูล", 13, 17, ORANGE)
    blocks = [
        ("M5:Q7", "ทำไมต้องปรับสเกล x?\nHours อยู่ประมาณ 1–9 แต่ Attendance อยู่ประมาณ 55–95 การทำให้ทั้งสองฟีเจอร์มีขนาดใกล้กันช่วยให้ Gradient Descent เสถียร และค่าน้ำหนักไม่ถูกเปรียบเทียบผ่านหน่วยที่ต่างกัน", SKY),
        ("M8:Q10", "ทำไมไม่ปรับสเกล y?\nคะแนน y คงไว้ในหน่วยคะแนน 0–100 เพื่อให้ Prediction, Residual, MAE และ RMSE ตีความเป็น “คะแนน” ได้โดยตรง แม้จะปรับ y ได้ในงานจริง แต่ต้องแปลงกลับก่อนรายงานผล", PALE_PURPLE),
        ("M12:Q14", "Hours center = 5\nมาจากค่าเฉลี่ย Hours ของ Train = 5.00 เมื่ออ่าน 5 ชั่วโมง จะได้ x1 = 0", PALE_GREEN),
        ("M15:Q17", "Hours scale = 2.5\nStandard deviation จริง ≈ 2.35 จึงปัดเป็น 2.5: 2.5/5/7.5 ชั่วโมง → -1/0/+1", PALE_ORANGE),
        ("M18:Q20", "Attendance center = 75\nค่าเฉลี่ยจริง = 74.5% จึงปัดเป็น 75% เมื่อเข้าเรียน 75% จะได้ x2 = 0", PALE_GREEN),
        ("M21:Q23", "Attendance scale = 15\nเลือกเพื่อให้ 60/75/90% → -1/0/+1 และทำให้การคำนวณมือจำง่าย", PALE_ORANGE),
        ("M25:Q27", "คะแนนจริงมี Noise\nข้อมูลจำลองไม่ได้อยู่บนเส้นตรงสมบูรณ์ เพราะข้อมูลจริงมักมีปัจจัยอื่น เช่น พื้นฐานเดิม การพักผ่อน และความยากของข้อสอบ Residual จึงไม่ควรเป็นศูนย์ทุกคน", SKY),
    ]
    for cell_range, text_value, color in blocks:
        merged_block(ws, cell_range, text_value, color)

    section(ws, 29, "เปรียบเทียบค่าที่ใช้กับสถิติจริงของ Train", 13, 17)
    headers2 = {13: "รายการ", 14: "ค่าในไฟล์", 15: "ค่าสถิติจริง", 16: "เหตุผลที่เลือก"}
    for c, value in headers2.items():
        ws.cell(30, c, value)
    header_row(ws, 30, 13, 17)
    ws.merge_cells("P30:Q30")
    stats_rows = [
        (31, "Hours center", "=$K$6", "=AVERAGE($B$6:$B$25)", "ใช้ค่าเฉลี่ย Train"),
        (32, "Hours scale", "=$K$7", "=STDEVP($B$6:$B$25)", "ปัดจาก 2.35 เป็น 2.5"),
        (33, "Attendance center", "=$K$8", "=AVERAGE($C$6:$C$25)", "ปัดจาก 74.5 เป็น 75"),
        (34, "Attendance scale", "=$K$9", "=STDEVP($C$6:$C$25)", "ใช้ 15 เพื่อแปลง 60/75/90 เป็น -1/0/+1"),
    ]
    for r, label, chosen, actual, reason in stats_rows:
        ws.cell(r, 13, label)
        ws.cell(r, 14, chosen)
        ws.cell(r, 15, actual)
        ws.cell(r, 16, reason)
        for c in range(13, 18):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = fill(SKY if r % 2 else WHITE)
            cell.font = font(9, c == 13, GREEN if c in (14, 15) else NAVY)
        ws.merge_cells(start_row=r, start_column=16, end_row=r, end_column=17)
        ws.cell(r, 14).number_format = "0.00"
        ws.cell(r, 15).number_format = "0.00"
        ws.row_dimensions[r].height = 30
    merged_block(ws, "M36:Q37", "Center-and-Scale ในไฟล์นี้ใช้ค่าที่ใกล้สถิติจริงแต่ปัดให้คำนวณง่าย หากต้องการ z-score มาตรฐาน ให้ใช้ Mean และ Standard Deviation จริงโดยไม่ปัด", PALE_PURPLE)

    ws.auto_filter.ref = "A5:H35"
    ws["A37"] = "แหล่งข้อมูล"
    ws["B37"] = "Synthetic dataset สำหรับสอน Linear Regression (16 กรกฎาคม 2026)"
    ws["A37"].font = font(9, True, NAVY)
    ws["B37"].font = font(9, False, GRAY, True)
    ws.merge_cells("B37:H37")
    ws.print_title_rows = "1:5"
    ws.print_area = "A1:Q37"


def build_training(wb):
    ws = wb.create_sheet("02_Training_Detail")
    configure_sheet(ws)
    title(
        ws,
        "02 • ฝึก Linear Regression แบบ Step by Step",
        "Batch Gradient Descent 15 รอบ • ใช้ Loss = 0.5 × Mean Squared Error",
        11,
    )
    ws.freeze_panes = "A11"
    for col, width in {"A": 13, "B": 11, "C": 11, "D": 11, "E": 13, "F": 13, "G": 13, "H": 14, "I": 14, "J": 14, "K": 15}.items():
        ws.column_dimensions[col].width = width

    section(ws, 3, "ค่าตั้งต้นที่ปรับทดลองได้", 1, 6, ORANGE)
    settings = [(4, "Learning rate (α)", 0.4), (5, "จำนวนรอบ", 15), (6, "จำนวน Train", '=COUNTIF(\'01_Data\'!E6:E35,"Train")')]
    for r, label, value in settings:
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        ws.cell(r, 1).font = font(10, True, NAVY)
        ws.cell(r, 1).fill = fill(LIGHT_GRAY)
        ws.cell(r, 2).fill = fill(PALE_ORANGE if r in (4, 5) else WHITE)
        ws.cell(r, 2).font = font(10, True, BLUE if r in (4, 5) else GREEN)
        ws.cell(r, 1).border = ws.cell(r, 2).border = grid_border
    ws["B4"].number_format = "0.00"
    ws["B4"].comment = Comment("ลองปรับ 0.10, 0.40 และ 0.80 แล้วดูกราฟ Loss", "OpenAI")
    ws["C7"], ws["D7"], ws["E7"], ws["F7"], ws["G7"], ws["H7"] = "b เริ่มต้น", 0, "w1 เริ่มต้น", 0, "w2 เริ่มต้น", 0
    for c in range(3, 9):
        cell = ws.cell(7, c)
        cell.border = grid_border
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill(LIGHT_GRAY if c % 2 else PALE_ORANGE)
        cell.font = font(10, True, NAVY if c % 2 else BLUE)
    ws["J4"] = "ทำไมเริ่มที่ 0?"
    ws["K4"] = "เป็นจุดเริ่มต้นที่เป็นกลาง โมเดลจะค่อย ๆ เรียนรู้ bias b และ weights w1, w2 จาก Gradient"
    ws["J4"].font = font(9, True, NAVY)
    ws["K4"].font = font(9, False, NAVY)
    ws["J4"].fill = fill(LIGHT_GRAY)
    ws["K4"].fill = fill(SKY)
    ws["J4"].border = ws["K4"].border = grid_border
    ws["K4"].alignment = Alignment(wrap_text=True)

    dv = DataValidation(type="decimal", operator="between", formula1="0.01", formula2="1", allow_blank=False)
    dv.error = "Learning rate ต้องอยู่ระหว่าง 0.01 ถึง 1.00"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(ws["B4"])

    meta = []
    for iteration in range(1, 16):
        start = 11 + (iteration - 1) * 26
        param_row = start + 1
        head_row = start + 2
        first_data = start + 3
        last_data = first_data + 19
        summary_row = last_data + 1
        new_row = summary_row + 1
        meta.append((param_row, summary_row, new_row))
        section(ws, start, f"รอบที่ {iteration:02d} • พารามิเตอร์ปัจจุบัน → ŷ → residual → gradient → squared error → พารามิเตอร์ใหม่", 1, 11, TEAL if iteration % 2 else NAVY)
        ws.cell(param_row, 1, "Iteration")
        ws.cell(param_row, 2, iteration)
        ws.cell(param_row, 3, "b")
        ws.cell(param_row, 5, "w1")
        ws.cell(param_row, 7, "w2")
        if iteration == 1:
            ws.cell(param_row, 4, "=$D$7")
            ws.cell(param_row, 6, "=$F$7")
            ws.cell(param_row, 8, "=$H$7")
        else:
            prev_new = meta[-2][2]
            ws.cell(param_row, 4, f"=D{prev_new}")
            ws.cell(param_row, 6, f"=F{prev_new}")
            ws.cell(param_row, 8, f"=H{prev_new}")
        ws.merge_cells(start_row=param_row, start_column=9, end_row=param_row, end_column=11)
        ws.cell(param_row, 9, "ใช้ค่าน้ำหนักนี้คำนวณทุกแถวในรอบ")
        for c in range(1, 12):
            cell = ws.cell(param_row, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = fill(PALE_BLUE if c in (4, 6, 8) else LIGHT_GRAY)
            cell.font = font(9, c in (1, 3, 5, 7), NAVY)

        headers = ["ID", "x1", "x2", "Actual y", "Prediction ŷ", "Residual ŷ-y", "grad b = error", "grad w1 = error×x1", "grad w2 = error×x2", "Squared Error", "0.5×Error²"]
        for c, value in enumerate(headers, 1):
            ws.cell(head_row, c, value)
        header_row(ws, head_row, 1, 11)
        for idx, r in enumerate(range(first_data, last_data + 1)):
            data_row = 6 + idx
            ws.cell(r, 1, f"='01_Data'!A{data_row}")
            ws.cell(r, 2, f"='01_Data'!F{data_row}")
            ws.cell(r, 3, f"='01_Data'!G{data_row}")
            ws.cell(r, 4, f"='01_Data'!D{data_row}")
            ws.cell(r, 5, f"=$D${param_row}+$F${param_row}*B{r}+$H${param_row}*C{r}")
            ws.cell(r, 6, f"=E{r}-D{r}")
            ws.cell(r, 7, f"=F{r}")
            ws.cell(r, 8, f"=F{r}*B{r}")
            ws.cell(r, 9, f"=F{r}*C{r}")
            ws.cell(r, 10, f"=F{r}^2")
            ws.cell(r, 11, f"=0.5*J{r}")
            for c in range(1, 12):
                cell = ws.cell(r, c)
                cell.border = grid_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font(9, False, GREEN if c <= 4 else BLACK)
                cell.fill = fill(SKY if c <= 4 else (LIGHT_GRAY if c in (6, 7, 8, 9) else WHITE))
            for c in range(2, 12):
                ws.cell(r, c).number_format = "0.0000"

        ws.cell(summary_row, 1, "สรุป Gradient และ Error")
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)
        ws.cell(summary_row, 7, f"=AVERAGE(G{first_data}:G{last_data})")
        ws.cell(summary_row, 8, f"=AVERAGE(H{first_data}:H{last_data})")
        ws.cell(summary_row, 9, f"=AVERAGE(I{first_data}:I{last_data})")
        ws.cell(summary_row, 10, f"=AVERAGE(J{first_data}:J{last_data})")
        ws.cell(summary_row, 11, f"=AVERAGE(K{first_data}:K{last_data})")
        for c in range(1, 12):
            cell = ws.cell(summary_row, c)
            cell.border = grid_border
            cell.fill = fill(PALE_PURPLE)
            cell.font = font(9, True, PURPLE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c >= 7:
                cell.number_format = "0.0000"
        ws.cell(summary_row, 7).comment = Comment("Average residual = gradient ของ bias b", "OpenAI")
        ws.cell(summary_row, 10).comment = Comment("Mean Squared Error ของรอบนี้", "OpenAI")
        ws.cell(summary_row, 11).comment = Comment("Loss ที่ใช้ฝึก = 0.5 × MSE เพื่อให้ derivative ไม่มีตัวคูณ 2", "OpenAI")

        ws.cell(new_row, 1, "Update: พารามิเตอร์ใหม่ = ค่าเดิม - α × gradient")
        ws.merge_cells(start_row=new_row, start_column=1, end_row=new_row, end_column=2)
        ws.cell(new_row, 3, "b ใหม่")
        ws.cell(new_row, 4, f"=D{param_row}-$B$4*G{summary_row}")
        ws.cell(new_row, 5, "w1 ใหม่")
        ws.cell(new_row, 6, f"=F{param_row}-$B$4*H{summary_row}")
        ws.cell(new_row, 7, "w2 ใหม่")
        ws.cell(new_row, 8, f"=H{param_row}-$B$4*I{summary_row}")
        ws.merge_cells(start_row=new_row, start_column=9, end_row=new_row, end_column=11)
        ws.cell(new_row, 9, "ค่าชุดนี้ใช้ในรอบถัดไป")
        for c in range(1, 12):
            cell = ws.cell(new_row, c)
            cell.border = grid_border
            cell.fill = fill(PALE_GREEN if c in (4, 6, 8) else LIGHT_GRAY)
            cell.font = font(9, c in (1, 3, 5, 7), GREEN if c in (4, 6, 8) else NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c in (4, 6, 8):
                cell.number_format = "0.0000"

    ws.print_title_rows = "1:10"
    ws.print_area = "A1:K399"
    return meta


def build_model_path(wb, meta):
    ws = wb.create_sheet("03_Model_Path")
    configure_sheet(ws)
    title(ws, "03 • เส้นทางการเรียนรู้ของโมเดล", "ติดตาม Weight, Gradient, MSE และ Loss ตลอด 15 รอบ", 21)
    ws.freeze_panes = "A6"
    headers = ["Iteration", "b ก่อน", "w1 ก่อน", "w2 ก่อน", "0.5×MSE", "MSE", "grad b", "grad w1", "grad w2", "b หลัง", "w1 หลัง", "w2 หลัง"]
    for c, value in enumerate(headers, 1):
        ws.cell(5, c, value)
        ws.column_dimensions[get_column_letter(c)].width = 13
    header_row(ws, 5, 1, 12)
    for r, (param_row, summary_row, new_row) in enumerate(meta, 6):
        ws.cell(r, 1, r - 5)
        refs = [
            f"='02_Training_Detail'!D{param_row}", f"='02_Training_Detail'!F{param_row}", f"='02_Training_Detail'!H{param_row}",
            f"='02_Training_Detail'!K{summary_row}", f"='02_Training_Detail'!J{summary_row}", f"='02_Training_Detail'!G{summary_row}",
            f"='02_Training_Detail'!H{summary_row}", f"='02_Training_Detail'!I{summary_row}", f"='02_Training_Detail'!D{new_row}",
            f"='02_Training_Detail'!F{new_row}", f"='02_Training_Detail'!H{new_row}",
        ]
        for c, formula in enumerate(refs, 2):
            ws.cell(r, c, formula)
        for c in range(1, 13):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center")
            cell.font = font(9, False, GREEN if c > 1 else NAVY)
            cell.fill = fill(SKY if r % 2 == 0 else WHITE)
            if c > 1:
                cell.number_format = "0.0000"
        ws.row_dimensions[r].height = 23

    section(ws, 23, "โมเดลสุดท้าย", 1, 6, ORANGE)
    for r, label, formula in [(24, "Final b", "=J20"), (25, "Final w1", "=K20"), (26, "Final w2", "=L20")]:
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula)
        ws.cell(r, 1).font = font(10, True, NAVY)
        ws.cell(r, 1).fill = fill(LIGHT_GRAY)
        ws.cell(r, 2).font = font(10, True, BLACK)
        ws.cell(r, 2).fill = fill(PALE_GREEN)
        ws.cell(r, 1).border = ws.cell(r, 2).border = grid_border
        ws.cell(r, 2).number_format = "0.0000"
    ws.merge_cells("D24:L24")
    ws["D24"] = "สมการบนข้อมูลที่ปรับสเกล: ŷ = w1x1 + w2x2 + b"
    ws["D24"].font = font(11, True, PURPLE)
    ws["D24"].fill = fill(PALE_PURPLE)
    ws["D24"].border = grid_border
    ws.merge_cells("D25:L26")
    ws["D25"] = "การตีความ: เมื่อ x1 เพิ่ม 1 หน่วย โดย x2 คงที่ คะแนนที่ทำนายเปลี่ยนประมาณ w1 คะแนน ส่วน w2 ตีความแบบเดียวกันกับ Attendance"
    ws["D25"].font = font(10, False, NAVY)
    ws["D25"].fill = fill(SKY)
    ws["D25"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["D25"].border = grid_border

    loss = LineChart()
    loss.title = "Training Loss ลดลงในแต่ละรอบ"
    loss.y_axis.title = "0.5 × MSE"
    loss.x_axis.title = "Iteration"
    loss.style = 13
    loss.height = 7.5
    loss.width = 15
    loss.add_data(Reference(ws, min_col=5, min_row=5, max_row=20), titles_from_data=True)
    loss.set_categories(Reference(ws, min_col=1, min_row=6, max_row=20))
    loss.legend = None
    ws.add_chart(loss, "N5")
    weights = LineChart()
    weights.title = "การเปลี่ยนแปลงของ b, w1, w2"
    weights.y_axis.title = "Weight"
    weights.x_axis.title = "Iteration"
    weights.style = 12
    weights.height = 7.5
    weights.width = 15
    weights.add_data(Reference(ws, min_col=10, max_col=12, min_row=5, max_row=20), titles_from_data=True)
    weights.set_categories(Reference(ws, min_col=1, min_row=6, max_row=20))
    ws.add_chart(weights, "N21")
    ws.print_area = "A1:U36"


def build_prediction(wb):
    ws = wb.create_sheet("04_Test_Prediction")
    configure_sheet(ws)
    title(ws, "04 • ทำนายข้อมูล Test และวิเคราะห์ Residual", "แยก contribution ของแต่ละ coefficient และคำนวณ error รายคน", 25)
    ws.freeze_panes = "A8"
    labels = [("A4", "Final b", "B4", "='03_Model_Path'!B24"), ("D4", "Final w1", "E4", "='03_Model_Path'!B25"), ("G4", "Final w2", "H4", "='03_Model_Path'!B26")]
    for label_cell, label, value_cell, formula in labels:
        ws[label_cell] = label
        ws[value_cell] = formula
        ws[label_cell].font = font(10, True, NAVY)
        ws[label_cell].fill = fill(LIGHT_GRAY)
        ws[value_cell].font = font(10, True, GREEN)
        ws[value_cell].fill = fill(SKY)
        ws[label_cell].border = ws[value_cell].border = grid_border
        ws[value_cell].number_format = "0.0000"
    headers = ["ID", "Hours", "Attend %", "x1", "x2", "Actual y", "b", "w1×x1", "w2×x2", "Prediction ŷ", "Residual ŷ-y", "Absolute Error", "Squared Error", "y-Mean(y)", "Squared Deviation"]
    for c, value in enumerate(headers, 1):
        ws.cell(7, c, value)
        ws.column_dimensions[get_column_letter(c)].width = [11, 10, 11, 11, 11, 11, 11, 12, 12, 14, 14, 14, 14, 14, 16][c - 1]
    header_row(ws, 7, 1, 15)
    for idx, r in enumerate(range(8, 18)):
        data_row = 26 + idx
        ws.cell(r, 1, f"='01_Data'!A{data_row}")
        ws.cell(r, 2, f"='01_Data'!B{data_row}")
        ws.cell(r, 3, f"='01_Data'!C{data_row}")
        ws.cell(r, 4, f"='01_Data'!F{data_row}")
        ws.cell(r, 5, f"='01_Data'!G{data_row}")
        ws.cell(r, 6, f"='01_Data'!D{data_row}")
        ws.cell(r, 7, "=$B$4")
        ws.cell(r, 8, f"=$E$4*D{r}")
        ws.cell(r, 9, f"=$H$4*E{r}")
        ws.cell(r, 10, f"=SUM(G{r}:I{r})")
        ws.cell(r, 11, f"=J{r}-F{r}")
        ws.cell(r, 12, f"=ABS(K{r})")
        ws.cell(r, 13, f"=K{r}^2")
        ws.cell(r, 14, f"=F{r}-AVERAGE($F$8:$F$17)")
        ws.cell(r, 15, f"=N{r}^2")
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = font(9, False, GREEN if c <= 6 else BLACK)
            cell.fill = fill(SKY if c <= 6 else WHITE)
        for c in range(4, 16):
            ws.cell(r, c).number_format = "0.0000"
        ws.row_dimensions[r].height = 26
    ws.conditional_formatting.add("L8:L17", CellIsRule(operator="lessThanOrEqual", formula=[2], fill=fill(PALE_GREEN)))
    ws.conditional_formatting.add("L8:L17", CellIsRule(operator="greaterThan", formula=[2], fill=fill(PALE_RED)))
    ws.auto_filter.ref = "A7:O17"

    compare = LineChart()
    compare.title = "Actual เทียบกับ Prediction"
    compare.y_axis.title = "คะแนน"
    compare.x_axis.title = "Student ID"
    compare.style = 13
    compare.height = 8
    compare.width = 16
    compare.add_data(Reference(ws, min_col=6, min_row=7, max_row=17), titles_from_data=True)
    compare.add_data(Reference(ws, min_col=10, min_row=7, max_row=17), titles_from_data=True)
    compare.set_categories(Reference(ws, min_col=1, min_row=8, max_row=17))
    ws.add_chart(compare, "Q7")
    residual = BarChart()
    residual.title = "Residual รายคน"
    residual.y_axis.title = "Prediction - Actual"
    residual.x_axis.title = "Student ID"
    residual.style = 12
    residual.height = 8
    residual.width = 16
    residual.add_data(Reference(ws, min_col=11, min_row=7, max_row=17), titles_from_data=True)
    residual.set_categories(Reference(ws, min_col=1, min_row=8, max_row=17))
    ws.add_chart(residual, "Q22")
    note_box(ws, 20, 1, 15, "Residual > 0 หมายถึงโมเดลทำนายสูงกว่าคะแนนจริง; Residual < 0 หมายถึงทำนายต่ำกว่าคะแนนจริง; ค่าที่ใกล้ 0 คือทำนายใกล้เคียง", SKY)
    ws.print_area = "A1:Y37"


def build_evaluation(wb):
    ws = wb.create_sheet("05_Evaluation")
    configure_sheet(ws)
    title(ws, "05 • ประเมินประสิทธิภาพ Linear Regression", "MAE, MSE, RMSE, R² และ Mean Residual จากข้อมูล Test", 10)
    for col, width in {"A": 4, "B": 22, "C": 18, "D": 25, "E": 25, "F": 4, "G": 24, "H": 24, "I": 24, "J": 4}.items():
        ws.column_dimensions[col].width = width
    section(ws, 4, "ค่ารวมที่ใช้คำนวณ", 2, 9)
    totals = [
        (5, "n (จำนวน Test)", "=COUNT('04_Test_Prediction'!F8:F17)", "จำนวนตัวอย่างที่ใช้วัดผล"),
        (6, "SSE", "=SUM('04_Test_Prediction'!M8:M17)", "ผลรวม Squared Error ของโมเดล"),
        (7, "SST", "=SUM('04_Test_Prediction'!O8:O17)", "ความแปรปรวนรวมเมื่อเทียบกับการทายค่าเฉลี่ย"),
    ]
    for r, label, formula, desc in totals:
        ws.cell(r, 2, label)
        ws.cell(r, 3, formula)
        ws.cell(r, 4, desc)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        for c in range(2, 10):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = fill(SKY if r % 2 else WHITE)
            cell.font = font(10, c == 2, GREEN if c == 3 else NAVY)
        ws.cell(r, 3).number_format = "0.0000"
        ws.row_dimensions[r].height = 28

    section(ws, 9, "Metrics", 2, 9, ORANGE)
    headers = {2: "Metric", 3: "ค่า", 4: "สูตร", 7: "ความหมาย"}
    for c, value in headers.items():
        ws.cell(10, c, value)
    header_row(ws, 10, 2, 9)
    ws.merge_cells("D10:F10")
    ws.merge_cells("G10:I10")
    metrics = [
        (11, "MAE", "=AVERAGE('04_Test_Prediction'!L8:L17)", "AVERAGE(|ŷ-y|)", "ผิดเฉลี่ยกี่คะแนน โดยให้น้ำหนักทุก error เท่ากัน"),
        (12, "MSE", "=AVERAGE('04_Test_Prediction'!M8:M17)", "AVERAGE((ŷ-y)²)", "ยกกำลังสองจึงลงโทษ error ขนาดใหญ่มากขึ้น"),
        (13, "RMSE", "=SQRT(C12)", "SQRT(MSE)", "ขนาด error ในหน่วยคะแนน และไวต่อ error ใหญ่กว่า MAE"),
        (14, "R²", "=1-C6/C7", "1-SSE/SST", "สัดส่วนความแปรปรวนของคะแนนที่โมเดลอธิบายได้; 1 คือสมบูรณ์"),
        (15, "Mean Residual", "=AVERAGE('04_Test_Prediction'!K8:K17)", "AVERAGE(ŷ-y)", "ตรวจ bias โดยรวม: บวกคือทำนายสูงไป ลบคือทำนายต่ำไป"),
    ]
    for r, label, formula, formula_text, meaning in metrics:
        ws.cell(r, 2, label)
        ws.cell(r, 3, formula)
        ws.cell(r, 4, formula_text)
        ws.cell(r, 7, meaning)
        for c in range(2, 10):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = fill(SKY if r % 2 else WHITE)
            cell.font = font(10, c == 2, TEAL if c == 3 else NAVY)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        ws.cell(r, 3).number_format = "0.0%" if label == "R²" else "0.000"
        ws.cell(r, 3).font = font(12, True, TEAL)
        ws.row_dimensions[r].height = 38
    note_box(ws, 18, 2, 9, '="ผล Test: MAE "&TEXT(C11,"0.00")&" คะแนน | RMSE "&TEXT(C13,"0.00")&" คะแนน | R² "&TEXT(C14,"0.0%")', PALE_GREEN)
    ws["B18"] = '="ผล Test: MAE "&TEXT(C11,"0.00")&" คะแนน | RMSE "&TEXT(C13,"0.00")&" คะแนน | R² "&TEXT(C14,"0.0%")'
    ws["B18"].font = font(11, True, GREEN)
    ws["B18"].fill = fill(PALE_GREEN)
    ws["B18"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["B18"].border = Border(left=medium_navy, right=thin_gray, top=thin_gray, bottom=thin_gray)
    merged_block(ws, "B21:I23", "ข้อควรระวัง: ค่า R² สูงไม่ได้ยืนยันว่าโมเดลมีเหตุและผลหรือเหมาะกับข้อมูลใหม่เสมอ ต้องตรวจ Train/Test split, residual pattern, outlier และความสมเหตุสมผลของข้อมูลด้วย", PALE_PURPLE)
    ws.print_area = "A1:J24"


def build_key_concepts(wb):
    ws = wb.create_sheet("06_Key_Concepts")
    configure_sheet(ws)
    title(ws, "06 • แนวคิดสำคัญที่ต้องเข้าใจ", "คำอธิบายว่า Linear Regression ทำอะไร ทำไปทำไม และควรตีความอย่างไร", 10)
    ws.freeze_panes = "A8"
    for col, width in {"A": 9, "B": 17, "C": 17, "D": 20, "E": 20, "F": 20, "G": 21, "H": 21, "I": 21, "J": 21}.items():
        ws.column_dimensions[col].width = width
    note_box(ws, 4, 1, 10, "ภาพรวม: โมเดลหาเส้นหรือระนาบที่ทำให้ผลรวมความคลาดเคลื่อนกำลังสองต่ำ โดยใช้ข้อมูล Train เรียนรู้ค่าน้ำหนัก แล้วใช้ Test ตรวจความสามารถกับข้อมูลที่ไม่เคยใช้ฝึก", SKY)
    section(ws, 7, "คำอธิบายแต่ละแนวคิด", 1, 10)
    for c, value in {1: "ขั้น", 2: "แนวคิด", 4: "ความหมาย", 7: "ทำไมสำคัญ / ข้อควรจำ"}.items():
        ws.cell(8, c, value)
    header_row(ws, 8, 1, 10)
    ws.merge_cells("B8:C8")
    ws.merge_cells("D8:F8")
    ws.merge_cells("G8:J8")
    concepts = [
        ("1", "Continuous target", "Linear Regression ใช้ทำนายค่าต่อเนื่อง เช่น คะแนน ราคา หรืออุณหภูมิ", "ถ้าเป้าหมายเป็น class 0/1 มักใช้ Logistic Regression แทน"),
        ("2", "Train / Test split", "Train ใช้เรียนรู้ weight ส่วน Test ใช้ประเมินหลังฝึก", "ช่วยตรวจว่าโมเดลใช้กับข้อมูลใหม่ได้ ไม่ใช่จำข้อมูล Train"),
        ("3", "Feature scaling", "ทำให้ฟีเจอร์ที่มีหน่วยต่างกันมีขนาดใกล้กัน", "ช่วยให้ Gradient Descent เสถียรและ coefficient หลัง scaling เปรียบเทียบง่ายขึ้น"),
        ("4", "Bias/Intercept b", "ค่าทำนายเมื่อ x1 และ x2 เท่ากับ 0", "เพราะ center ทำให้ x=0 แทนค่าทั่วไป b จึงใกล้คะแนนของนักศึกษาทั่วไป"),
        ("5", "Weights w1/w2", "คะแนนที่คาดว่าจะเปลี่ยนเมื่อฟีเจอร์เพิ่ม 1 scaled unit โดยอีกฟีเจอร์คงที่", "เป็น association ภายในข้อมูล ไม่ได้ยืนยันเหตุและผลโดยอัตโนมัติ"),
        ("6", "Prediction ŷ", "ค่าที่โมเดลประมาณจาก ŷ = w1x1 + w2x2 + b", "ŷ เป็นค่าต่อเนื่องและอาจเกินช่วงข้อมูลได้ เพราะไม่มี Sigmoid จำกัดค่า"),
        ("7", "Residual", "ŷ-y; ระยะห่างพร้อมเครื่องหมายระหว่างค่าทำนายกับค่าจริง", "ควรกระจายรอบ 0 โดยไม่มี pattern ชัดเจน"),
        ("8", "Squared Error", "Residual² ทำให้ error ทุกค่าบวกและลงโทษ error ใหญ่", "Outlier มีอิทธิพลสูงเพราะ error ถูกยกกำลังสอง"),
        ("9", "0.5×MSE Loss", "ค่าเฉลี่ย 0.5×Residual² ที่ใช้ฝึก", "ตัวคูณ 0.5 ทำให้ derivative ง่ายขึ้น แต่ตำแหน่งค่าต่ำสุดไม่เปลี่ยน"),
        ("10", "Gradient", "บอกทิศทางและขนาดที่ควรปรับ coefficient", "เมื่อ gradient เข้าใกล้ 0 แปลว่าอยู่ใกล้จุดต่ำสุดของ Loss"),
        ("11", "Learning rate", "กำหนดขนาดก้าวของการอัปเดต weight", "เล็กเกินไปเรียนช้า ใหญ่เกินไปอาจแกว่งหรือ Loss เพิ่ม"),
        ("12", "MAE", "ค่าเฉลี่ย Absolute Error", "ตีความง่ายเป็นคะแนนและไม่ลงโทษ error ใหญ่แรงเท่า RMSE"),
        ("13", "MSE / RMSE", "MSE เฉลี่ย error²; RMSE ถอดรากให้กลับเป็นหน่วยคะแนน", "เหมาะเมื่ออยากลงโทษ error ขนาดใหญ่มากขึ้น"),
        ("14", "R²", "เทียบโมเดลกับ baseline ที่ทายค่าเฉลี่ยทุกคน", "R² อาจติดลบได้หากโมเดลแย่กว่าการทายค่าเฉลี่ย"),
        ("15", "Overfitting", "โมเดลทำดีบน Train แต่แย่บนข้อมูลใหม่", "ต้องดู Test metrics และ residual ไม่ใช่ Training Loss อย่างเดียว"),
    ]
    for r, (step, concept, meaning, importance) in enumerate(concepts, 9):
        ws.cell(r, 1, step)
        ws.cell(r, 2, concept)
        ws.cell(r, 4, meaning)
        ws.cell(r, 7, importance)
        for c in range(1, 11):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center" if c == 1 else "left", vertical="center", wrap_text=True)
            cell.fill = fill(SKY if r % 2 else WHITE)
            cell.font = font(9, c in (1, 2), TEAL if c in (1, 2) else NAVY)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)
        ws.row_dimensions[r].height = 52
    section(ws, 25, "แปลง coefficient จาก scaled unit เป็นหน่วยดิบ", 1, 10, ORANGE)
    merged_block(ws, "A26:J28", "เพราะ x1=(Hours-5)/2.5 ผลของ Hours ต่อ 1 ชั่วโมง ≈ w1/2.5 คะแนน และผลของ Attendance ต่อ 1 จุดเปอร์เซ็นต์ ≈ w2/15 คะแนน ส่วน bias/intercept b ในหน่วยดิบต้องปรับ center กลับด้วย", PALE_ORANGE)
    note_box(ws, 30, 1, 10, "การตีความ coefficient ต้องสมมติว่าอีกฟีเจอร์คงที่ และสะท้อนความสัมพันธ์ในข้อมูลชุดนี้ ไม่ใช่หลักฐานเชิงสาเหตุ", PALE_PURPLE)
    ws.print_area = "A1:J32"


def build_exercises(wb):
    ws = wb.create_sheet("07_Exercises")
    configure_sheet(ws)
    title(ws, "07 • แบบฝึกหัดและทดลองโมเดล", "ทำนายนักศึกษาใหม่ ทดลอง Learning rate และอภิปรายการตีความ", 12)
    for col, width in {"A": 20, "B": 16, "C": 20, "D": 18, "E": 18, "F": 18, "G": 18, "H": 4, "I": 24, "J": 24, "K": 24, "L": 24}.items():
        ws.column_dimensions[col].width = width
    section(ws, 4, "ทดลองทำนายนักศึกษาใหม่", 1, 7, ORANGE)
    inputs = [(5, "Hours", 4.0), (6, "Attendance (%)", 80.0)]
    for r, label, value in inputs:
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        ws.cell(r, 1).font = font(10, True, NAVY)
        ws.cell(r, 1).fill = fill(LIGHT_GRAY)
        ws.cell(r, 2).font = font(10, True, BLUE)
        ws.cell(r, 2).fill = fill(PALE_ORANGE)
        ws.cell(r, 1).border = ws.cell(r, 2).border = grid_border
    ws["A8"], ws["B8"] = "x1", "=(B5-'01_Data'!K6)/'01_Data'!K7"
    ws["A9"], ws["B9"] = "x2", "=(B6-'01_Data'!K8)/'01_Data'!K9"
    ws["A11"], ws["B11"] = "Predicted score", "='03_Model_Path'!B24+'03_Model_Path'!B25*B8+'03_Model_Path'!B26*B9"
    for r in (8, 9, 11):
        ws.cell(r, 1).font = font(10, True, NAVY)
        ws.cell(r, 1).fill = fill(LIGHT_GRAY)
        ws.cell(r, 2).font = font(11, True, GREEN)
        ws.cell(r, 2).fill = fill(PALE_GREEN if r == 11 else SKY)
        ws.cell(r, 1).border = ws.cell(r, 2).border = grid_border
        ws.cell(r, 2).number_format = "0.00"
    ws.merge_cells("D5:G11")
    ws["D5"] = "ขั้นตอน\n1) กรอก Hours และ Attendance ในช่องสีเหลือง\n2) Excel ปรับสเกลเป็น x1, x2\n3) นำค่าน้ำหนักสุดท้ายมาคำนวณ ŷ\n4) ทดลองเปลี่ยนข้อมูลและอธิบายว่าทำไมคะแนนเปลี่ยน"
    ws["D5"].font = font(10, True, NAVY)
    ws["D5"].fill = fill(SKY)
    ws["D5"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["D5"].border = Border(left=medium_navy, right=thin_gray, top=thin_gray, bottom=thin_gray)

    section(ws, 14, "บันทึกผลเมื่อเปลี่ยน Learning rate", 1, 7)
    for c, value in enumerate(["Learning rate", "Final Loss", "MAE", "RMSE", "R²", "Loss ลดต่อเนื่อง?", "ข้อสังเกต"], 1):
        ws.cell(15, c, value)
    header_row(ws, 15, 1, 7)
    for r, alpha in enumerate([0.10, 0.40, 0.80], 16):
        ws.cell(r, 1, alpha)
        for c in range(1, 8):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = fill(PALE_ORANGE if c > 1 else SKY)
            cell.font = font(10, c == 1, BLUE if c == 1 else GRAY)
        ws.cell(r, 1).number_format = "0.00"
        ws.row_dimensions[r].height = 34
    merged_block(ws, "I4:L11", "วิธีทดลอง Learning rate\nกลับไปแก้ 02_Training_Detail!B4 แล้วจด Final Loss และ Test metrics ลงตาราง เปรียบเทียบว่า α เล็กทำให้เรียนช้าเพียงใด และ α ใหญ่ทำให้ Loss แกว่งหรือไม่", PALE_PURPLE)

    section(ws, 21, "คำถามสำหรับอภิปราย", 1, 12, ORANGE)
    questions = [
        "1. b, w1 และ w2 มีเครื่องหมายอย่างไร และสอดคล้องกับบริบทหรือไม่?",
        "2. MAE กับ RMSE ต่างกันอย่างไร และเหตุใด RMSE มักสูงกว่า MAE?",
        "3. นักศึกษาคนใดมี Absolute Error สูงที่สุด? มีลักษณะข้อมูลอย่างไร?",
        "4. หากเพิ่มฟีเจอร์คุณภาพการนอน โมเดลอาจดีขึ้นอย่างไร และต้องระวังอะไร?",
        "5. เหตุใด R² สูงจึงไม่ใช่หลักฐานว่าการอ่านหนังสือเป็นสาเหตุโดยตรงของคะแนน?",
    ]
    for r, question in enumerate(questions, 22):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.cell(r, 1, question)
        ws.cell(r, 1).font = font(10, r == 22, NAVY)
        ws.cell(r, 1).fill = fill(SKY if r % 2 else WHITE)
        ws.cell(r, 1).border = grid_border
        ws.cell(r, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
        ws.cell(r, 9, "บันทึกคำตอบ:")
        ws.cell(r, 9).font = font(9, False, GRAY, True)
        ws.cell(r, 9).fill = fill(PALE_ORANGE)
        ws.cell(r, 9).border = grid_border
        ws.row_dimensions[r].height = 36
    ws.print_area = "A1:L28"


def build_workbook():
    wb = Workbook()
    wb.properties.title = "Linear Regression Step by Step for Students"
    wb.properties.subject = "Excel-based Linear Regression teaching workbook"
    wb.properties.creator = "OpenAI"
    wb.properties.description = "Thai classroom workbook with formula-driven gradient descent, residual analysis, and regression metrics"
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb._named_styles["Normal"].font = font()
    build_guide(wb)
    build_data(wb)
    meta = build_training(wb)
    build_model_path(wb, meta)
    build_prediction(wb)
    build_evaluation(wb)
    build_key_concepts(wb)
    build_exercises(wb)
    wb.active = 0
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    build_workbook()
