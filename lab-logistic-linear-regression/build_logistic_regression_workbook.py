import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path(__file__).with_name("Logistic_Regression_Step_by_Step_TH.xlsx")
TRAINING_ITERATIONS = 15
PRACTICE_MODE = False

NAVY = "17324D"
TEAL = "0B7285"
BLUE = "1971C2"
SKY = "E7F5FF"
PALE_BLUE = "D0EBFF"
ORANGE = "F59F00"
PALE_ORANGE = "FFF3BF"
GREEN = "2B8A3E"
PALE_GREEN = "D3F9D8"
RED = "C92A2A"
PALE_RED = "FFE3E3"
PURPLE = "7048E8"
PALE_PURPLE = "E5DBFF"
GRAY = "495057"
LIGHT_GRAY = "F1F3F5"
MID_GRAY = "CED4DA"
WHITE = "FFFFFF"
BLACK = "000000"

thin_gray = Side(style="thin", color=MID_GRAY)
medium_navy = Side(style="medium", color=NAVY)
grid_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)


def fill(color):
    return PatternFill("solid", fgColor=color)


def font(size=10, bold=False, color=BLACK, italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)


def title(ws, text, subtitle, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1, text)
    ws.cell(1, 1).font = font(20, True, WHITE)
    ws.cell(1, 1).fill = fill(NAVY)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = font(10, False, WHITE)
    ws.cell(2, 1).fill = fill(TEAL)
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 25


def section(ws, row, text, start_col, end_col, color=TEAL):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row, start_col, text)
    c.font = font(11, True, WHITE)
    c.fill = fill(color)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 23


def header_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        c = ws.cell(row, col)
        c.font = font(9, True, WHITE)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = grid_border
    ws.row_dimensions[row].height = 34


def style_grid(ws, min_row, max_row, min_col, max_col, center=False):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for c in row:
            c.border = grid_border
            c.font = font()
            c.alignment = Alignment(
                horizontal="center" if center else "left",
                vertical="center",
                wrap_text=True,
            )


def note_box(ws, row, start_col, end_col, text, color=SKY):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row + 1, end_column=end_col)
    c = ws.cell(row, start_col, text)
    c.fill = fill(color)
    c.font = font(10, False, NAVY)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = Border(left=medium_navy, right=thin_gray, top=thin_gray, bottom=thin_gray)
    ws.row_dimensions[row].height = 25
    ws.row_dimensions[row + 1].height = 25


def configure_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.oddFooter.center.text = "Logistic Regression • Step by Step"
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.color = GRAY


def build_guide(wb):
    ws = wb.active
    ws.title = "00_Guide"
    configure_sheet(ws)
    title(
        ws,
        "Logistic Regression — แบบฝึกคำนวณทีละขั้นใน Excel" if PRACTICE_MODE else "Logistic Regression — คำนวณทีละขั้นใน Excel",
        (
            "แบบฝึกก่อนสอบ: เติมช่องสีเหลืองให้ครบ 20 รอบ แล้วตรวจเส้นทางการเรียนรู้ของโมเดล"
            if PRACTICE_MODE
            else "จากข้อมูลดิบ → Gradient Descent → Probability → Prediction → Confusion Matrix → Metrics"
        ),
        8,
    )
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 4

    section(ws, 4, "เป้าหมายการเรียนรู้", 2, 7)
    goals = [
        "1. อธิบายสมการ z = w1x1 + w2x2 + b และ sigmoid ได้",
        "2. เห็นว่า Gradient Descent ปรับค่าน้ำหนักอย่างไรในแต่ละรอบ",
        "3. เปลี่ยน probability เป็น class ด้วย threshold และอธิบาย TP/FP/FN/TN ได้",
        "4. คำนวณ Accuracy, Precision, Recall และ F1-score จาก Confusion Matrix ได้",
    ]
    for i, text in enumerate(goals, 5):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        c = ws.cell(i, 2, text)
        c.fill = fill(SKY if i % 2 else LIGHT_GRAY)
        c.font = font(10, False, NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = grid_border
        ws.row_dimensions[i].height = 28

    section(ws, 10, "เส้นทางการเรียนใน Workbook", 2, 7)
    steps = [
        ("01_Data", "ข้อมูลนักศึกษา 30 คน พร้อมการปรับสเกล x1 และ x2"),
        ("02_Training_Detail", "คำนวณ z, p, error, gradient, log loss และอัปเดต w1, w2, b ทีละรอบ"),
        ("03_Model_Path", f"ดูตารางสรุปค่าน้ำหนักและกราฟ loss ตลอด {TRAINING_ITERATIONS} รอบ"),
        ("04_Test_Prediction", "ใช้โมเดลกับข้อมูล Test และทดลองเปลี่ยน threshold"),
        ("05_Evaluation", "สร้าง Confusion Matrix และคำนวณ metrics จากสูตร"),
        ("06_Key_Concepts", "อ่านคำอธิบายแนวคิดสำคัญและเหตุผลของแต่ละขั้นตอน"),
        ("07_Exercises", "ทดลอง threshold หลายค่าและตอบคำถามอภิปราย"),
    ]
    for i, (name, desc) in enumerate(steps, 11):
        ws.cell(i, 2, name)
        ws.cell(i, 2).font = font(10, True, TEAL)
        ws.cell(i, 2).fill = fill(PALE_BLUE)
        ws.cell(i, 2).border = grid_border
        ws.cell(i, 2).alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
        ws.cell(i, 3, desc)
        ws.cell(i, 3).font = font()
        ws.cell(i, 3).fill = fill(WHITE)
        ws.cell(i, 3).border = grid_border
        ws.cell(i, 3).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 26

    section(ws, 18, "สูตรหลัก", 2, 7)
    formulas = [
        ("Linear score", "z = w1x1 + w2x2 + b"),
        ("Sigmoid", "p = 1 / (1 + EXP(-z))"),
        ("Error", "error = p - y"),
        ("Gradient", "grad b = AVERAGE(error); grad w1 = AVERAGE(error×x1); grad w2 = AVERAGE(error×x2)"),
        ("Update", "พารามิเตอร์ใหม่ = พารามิเตอร์เดิม - learning rate × gradient"),
        ("Log Loss", "-[y·LN(p) + (1-y)·LN(1-p)]"),
    ]
    for i, (label, formula_text) in enumerate(formulas, 19):
        ws.cell(i, 2, label)
        ws.cell(i, 2).font = font(10, True, NAVY)
        ws.cell(i, 2).fill = fill(LIGHT_GRAY)
        ws.cell(i, 2).border = grid_border
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=7)
        ws.cell(i, 3, formula_text)
        ws.cell(i, 3).font = Font(name="Arial", size=10, color=PURPLE)
        ws.cell(i, 3).fill = fill(PALE_PURPLE)
        ws.cell(i, 3).border = grid_border
        ws.cell(i, 3).alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 25

    section(ws, 26, "รหัสสีและวิธีทดลอง", 2, 7)
    legends = [
        (PALE_ORANGE, BLUE, "ช่องสีเหลือง = ช่องว่างที่นักศึกษาต้องเติม" if PRACTICE_MODE else "ช่องสีเหลือง = ค่าที่นักศึกษาปรับทดลองได้"),
        (WHITE, BLACK, "ตัวอักษรสีดำ = สูตรคำนวณภายในชีต"),
        (WHITE, GREEN, "ตัวอักษรสีเขียว = สูตรที่ดึงค่าจากชีตอื่น"),
        (PALE_GREEN, GREEN, "สีเขียว = ผลลัพธ์ถูกต้อง / TP / TN"),
        (PALE_RED, RED, "สีแดง = ผลลัพธ์คลาดเคลื่อน / FP / FN"),
    ]
    for i, (bg, fg, text) in enumerate(legends, 27):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        c = ws.cell(i, 2, text)
        c.fill = fill(bg)
        c.font = font(10, True if i == 27 else False, fg)
        c.border = grid_border
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 24

    note_box(
        ws,
        33,
        2,
        7,
        (
            "คำแนะนำ: เริ่มจาก 01_Data แล้วเติมช่องสีเหลืองใน 02_Training_Detail ตามลำดับรอบที่ 1–20 "
            "เมื่อกรอกครบ 03_Model_Path และชีตถัดไปจะแสดงผลจากคำตอบของนักศึกษา"
            if PRACTICE_MODE
            else "คำแนะนำ: เริ่มจากดูข้อมูลใน 01_Data แล้วเปิด 02_Training_Detail รอบที่ 1 ให้ครบหนึ่งรอบ "
            "จากนั้นใช้ 03_Model_Path สรุปภาพใหญ่ ก่อนทดลอง threshold ใน 04_Test_Prediction"
        ),
        PALE_ORANGE,
    )
    ws.print_area = "A1:H35"


def build_data(wb):
    ws = wb.create_sheet("01_Data")
    configure_sheet(ws)
    title(
        ws,
        "01 • ข้อมูลและการเตรียมฟีเจอร์",
        "ข้อมูลจำลองเพื่อการเรียนการสอน: 20 แถวแรกเป็น Train และ 10 แถวท้ายเป็น Test • ด้านขวามีคำอธิบายที่มาของการปรับสเกล",
        17,
    )
    ws.page_setup.orientation = "landscape"
    ws.freeze_panes = "A6"
    headers = [
        "Student ID",
        "ชั่วโมงอ่าน/วัน",
        "เข้าเรียน (%)",
        "ผลจริง y\n(1=ผ่าน)",
        "Split",
        "x1\n(Centered & scaled hours)",
        "x2\n(Centered & scaled attendance)",
        "คำอธิบาย",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col, value)
    header_row(ws, 5, 1, 8)
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 19
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 3
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 3
    ws.column_dimensions["M"].width = 21
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 22
    ws.column_dimensions["Q"].width = 22

    data = [
        ("S01", 1.0, 55, 0, "Train"), ("S02", 1.5, 60, 0, "Train"),
        ("S03", 2.0, 58, 0, "Train"), ("S04", 2.5, 65, 0, "Train"),
        ("S05", 3.0, 62, 0, "Train"), ("S06", 3.5, 70, 0, "Train"),
        ("S07", 4.0, 68, 0, "Train"), ("S08", 4.5, 72, 1, "Train"),
        ("S09", 5.0, 75, 1, "Train"), ("S10", 5.5, 78, 1, "Train"),
        ("S11", 6.0, 80, 1, "Train"), ("S12", 6.5, 82, 1, "Train"),
        ("S13", 7.0, 85, 1, "Train"), ("S14", 7.5, 88, 1, "Train"),
        ("S15", 8.0, 90, 1, "Train"), ("S16", 8.5, 92, 1, "Train"),
        ("S17", 9.0, 95, 1, "Train"), ("S18", 5.0, 65, 0, "Train"),
        ("S19", 3.0, 85, 1, "Train"), ("S20", 7.0, 65, 1, "Train"),
        ("T01", 2.0, 55, 0, "Test"), ("T02", 3.0, 75, 0, "Test"),
        ("T03", 4.0, 82, 1, "Test"), ("T04", 5.0, 65, 0, "Test"),
        ("T05", 5.0, 90, 1, "Test"), ("T06", 6.0, 72, 1, "Test"),
        ("T07", 7.0, 60, 0, "Test"), ("T08", 2.5, 95, 1, "Test"),
        ("T09", 8.0, 80, 1, "Test"), ("T10", 4.5, 70, 1, "Test"),
    ]
    for r, row in enumerate(data, 6):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
        ws.cell(r, 6, f"=(B{r}-$K$6)/$K$7")
        ws.cell(r, 7, f"=(C{r}-$K$8)/$K$9")
        ws.cell(r, 8, f'=IF(E{r}="Train","ใช้ฝึกโมเดล","เก็บไว้ประเมิน")')
        for c in range(1, 9):
            ws.cell(r, c).border = grid_border
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(1, 6):
            ws.cell(r, c).font = font(10, False, BLUE)
        for c in range(6, 9):
            ws.cell(r, c).font = font()
        ws.cell(r, 6).number_format = "0.000"
        ws.cell(r, 7).number_format = "0.000"
        ws.cell(r, 3).number_format = "0\"%\""
        ws.cell(r, 5).fill = fill(PALE_BLUE if row[4] == "Train" else PALE_ORANGE)
        ws.row_dimensions[r].height = 23

    section(ws, 4, "ข้อมูลหลัก", 1, 8)
    section(ws, 4, "ค่าที่ใช้ปรับสเกล", 10, 11, ORANGE)
    assumptions = [
        ("Hours center", 5.0),
        ("Hours scale", 2.5),
        ("Attendance center", 75.0),
        ("Attendance scale", 15.0),
    ]
    for r, (label, value) in enumerate(assumptions, 6):
        ws.cell(r, 10, label)
        ws.cell(r, 11, value)
        for c in (10, 11):
            ws.cell(r, c).border = grid_border
            ws.cell(r, c).alignment = Alignment(vertical="center")
        ws.cell(r, 10).fill = fill(LIGHT_GRAY)
        ws.cell(r, 10).font = font(9, True, NAVY)
        ws.cell(r, 11).fill = fill(PALE_ORANGE)
        ws.cell(r, 11).font = font(10, True, BLUE)
    ws["J11"] = "สูตร"
    ws["J11"].font = font(9, True, NAVY)
    ws["K11"] = "x = (ค่าดิบ - center) / scale"
    ws["K11"].font = font(9, False, PURPLE)
    ws["K11"].alignment = Alignment(wrap_text=True)
    ws["J11"].border = ws["K11"].border = grid_border
    ws["J11"].fill = fill(LIGHT_GRAY)
    ws["K11"].fill = fill(PALE_PURPLE)
    ws["K6"].comment = Comment("ค่าเฉลี่ยของชั่วโมงอ่านในชุด Train เท่ากับ 5.00 จึงใช้เป็นจุดกึ่งกลาง ทำให้ 5 ชั่วโมงแปลงเป็น x1 = 0", "OpenAI")
    ws["K7"].comment = Comment("ส่วนเบี่ยงเบนมาตรฐานจริงประมาณ 2.35 จึงปัดเป็น 2.5 เพื่อให้คำนวณมือและตีความง่าย ค่านี้เป็น teaching scale ไม่ใช่ z-score แบบเคร่งครัด", "OpenAI")
    ws["K8"].comment = Comment("ค่าเฉลี่ย Attendance ในชุด Train เท่ากับ 74.5% จึงปัดเป็น 75% เพื่อใช้เป็นจุดกึ่งกลางที่จำง่าย", "OpenAI")
    ws["K9"].comment = Comment("เลือก 15 จุดเปอร์เซ็นต์เพื่อให้ 60% แปลงเป็น -1, 75% เป็น 0 และ 90% เป็น +1 ค่านี้เน้นการสอนและการคำนวณมือ", "OpenAI")

    section(ws, 4, "ทำไมต้องปรับสเกล และตัวเลขมาจากไหน", 13, 17, ORANGE)
    explanation_blocks = [
        ("M5:Q7", "ทำไมต้องปรับสเกล?\nHours อยู่ประมาณ 1–9 แต่ Attendance อยู่ประมาณ 55–95 หากใช้ค่าดิบ Gradient ของ Attendance อาจมีขนาดใหญ่กว่าเพียงเพราะหน่วยต่างกัน การปรับสเกลทำให้ทั้งสองฟีเจอร์มีขนาดใกล้กัน เรียนรู้เสถียรขึ้น และใช้ learning rate เดียวกันได้เหมาะสมกว่า", SKY),
        ("M8:Q10", "สูตรที่ใช้\nx_scaled = (ค่าดิบ - center) / scale\nการลบ center ทำให้ค่าทั่วไปอยู่ใกล้ 0 ส่วนการหาร scale ทำให้ระยะห่างอยู่ราว -1 ถึง +1 โดยไฟล์นี้ใช้ Center-and-Scale แบบออกแบบเพื่อการสอน ไม่ใช่ z-score แบบเคร่งครัด", PALE_PURPLE),
        ("M12:Q14", "Hours center = 5 มาจากไหน?\nค่าเฉลี่ยชั่วโมงอ่านของข้อมูล Train เท่ากับ 5.00 ชั่วโมงพอดี จึงใช้ 5 เป็นจุดอ้างอิง เมื่ออ่าน 5 ชั่วโมง จะได้ x1 = (5-5)/2.5 = 0", PALE_GREEN),
        ("M15:Q17", "Hours scale = 2.5 มาจากไหน?\nส่วนเบี่ยงเบนมาตรฐานจริงของ Train ≈ 2.35 ชั่วโมง จึงปัดเป็น 2.5 เพื่อให้นักศึกษาคำนวณง่าย เช่น 2.5 ชม. → -1, 5 ชม. → 0, 7.5 ชม. → +1", PALE_ORANGE),
        ("M18:Q20", "Attendance center = 75 มาจากไหน?\nค่าเฉลี่ย Attendance ของ Train = 74.5% จึงปัดเป็น 75% ซึ่งเป็นค่ากลางที่จำง่าย เมื่อเข้าเรียน 75% จะได้ x2 = 0", PALE_GREEN),
        ("M21:Q23", "Attendance scale = 15 มาจากไหน?\nเลือก 15 จุดเปอร์เซ็นต์เพื่อให้ตีความง่าย: 60% → -1, 75% → 0, 90% → +1 และทำให้ช่วงข้อมูล Train 55–95% อยู่ราว -1.33 ถึง +1.33", PALE_ORANGE),
        ("M25:Q27", "ข้อควรจำ\nCenter กำหนดว่า “ค่าใดควรเป็นศูนย์” ส่วน Scale กำหนดว่า “การเปลี่ยนเท่าใดควรนับเป็นหนึ่งหน่วย” ค่าน้ำหนัก w1 และ w2 หลังการ scaling จึงเปรียบเทียบกันได้ง่ายขึ้น", SKY),
    ]
    for cell_range, text_value, bg in explanation_blocks:
        ws.merge_cells(cell_range)
        c = ws[cell_range.split(":")[0]]
        c.value = text_value
        c.font = font(9, True, NAVY)
        c.fill = fill(bg)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = Border(left=medium_navy, right=thin_gray, top=thin_gray, bottom=thin_gray)

    section(ws, 29, "เปรียบเทียบค่าที่ใช้กับสถิติจริงของ Train", 13, 17, TEAL)
    stats_headers = ["รายการ", "ค่าในไฟล์", "ค่าสถิติจริง", "เหตุผลที่เลือก"]
    ws["M30"], ws["N30"], ws["O30"], ws["P30"] = stats_headers
    ws.merge_cells("P30:Q30")
    header_row(ws, 30, 13, 17)
    stats_rows = [
        (31, "Hours center", "=$K$6", "=AVERAGE($B$6:$B$25)", "ใช้ค่าเฉลี่ยจริงของ Train"),
        (32, "Hours scale", "=$K$7", "=STDEVP($B$6:$B$25)", "ปัดจาก 2.35 เป็น 2.5 เพื่อคำนวณง่าย"),
        (33, "Attendance center", "=$K$8", "=AVERAGE($C$6:$C$25)", "ปัดจาก 74.5 เป็น 75 เพื่อจำง่าย"),
        (34, "Attendance scale", "=$K$9", "=STDEVP($C$6:$C$25)", "ใช้ 15 เพื่อให้ 60/75/90 แปลงเป็น -1/0/+1"),
    ]
    for r, label, selected, actual, reason in stats_rows:
        ws.cell(r, 13, label)
        ws.cell(r, 14, selected)
        ws.cell(r, 15, actual)
        ws.cell(r, 16, reason)
        ws.merge_cells(start_row=r, start_column=16, end_row=r, end_column=17)
        for col in range(13, 18):
            c = ws.cell(r, col)
            c.border = grid_border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.font = font(9, col == 13, GREEN if col in (14, 15) else NAVY)
            c.fill = fill(SKY if r % 2 else WHITE)
        ws.cell(r, 14).number_format = "0.00"
        ws.cell(r, 15).number_format = "0.00"
        ws.row_dimensions[r].height = 30
    ws.merge_cells("M36:Q37")
    ws["M36"] = "หากต้องการ z-score แบบมาตรฐาน ให้ใช้ center = AVERAGE(Train) และ scale = STDEV.P(Train) โดยตรง แต่ตัวอย่างนี้ตั้งใจปัดค่าเพื่อให้ติดตามสูตรด้วยมือได้ง่าย"
    ws["M36"].font = font(9, True, PURPLE)
    ws["M36"].fill = fill(PALE_PURPLE)
    ws["M36"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["M36"].border = Border(left=medium_navy, right=thin_gray, top=thin_gray, bottom=thin_gray)
    ws.auto_filter.ref = "A5:H35"
    ws.conditional_formatting.add("D6:D35", CellIsRule(operator="equal", formula=[1], fill=fill(PALE_GREEN)))
    ws.conditional_formatting.add("D6:D35", CellIsRule(operator="equal", formula=[0], fill=fill(LIGHT_GRAY)))
    ws["A37"] = "แหล่งข้อมูล"
    ws["B37"] = "Synthetic dataset สร้างเพื่อการเรียนการสอน Logistic Regression (15 กรกฎาคม 2026)"
    ws["A37"].font = font(9, True, NAVY)
    ws["B37"].font = font(9, False, GRAY, True)
    ws.merge_cells("B37:H37")
    ws.print_title_rows = "1:5"
    ws.print_area = "A1:Q37"


def build_training(wb):
    ws = wb.create_sheet("02_Training_Detail")
    configure_sheet(ws)
    title(
        ws,
        "02 • ฝึกโมเดลแบบ Step by Step",
        (
            f"แบบฝึก Batch Gradient Descent: เติมค่าที่คำนวณได้ด้วยตนเองให้ครบ {TRAINING_ITERATIONS} รอบ"
            if PRACTICE_MODE
            else f"Batch Gradient Descent: ทุกสูตรคำนวณใน Excel และเชื่อมต่อกัน {TRAINING_ITERATIONS} รอบ"
        ),
        11,
    )
    ws.freeze_panes = "A11"
    widths = {"A": 13, "B": 11, "C": 11, "D": 10, "E": 13, "F": 13, "G": 13, "H": 13, "I": 13, "J": 13, "K": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    section(ws, 3, "ค่าตั้งต้นที่ปรับทดลองได้", 1, 6, ORANGE)
    settings = [
        (4, "Learning rate (α)", 0.4),
        (5, "จำนวนรอบ", TRAINING_ITERATIONS),
        (6, "จำนวน Train", '=COUNTIF(\'01_Data\'!E6:E35,"Train")'),
    ]
    for r, label, value in settings:
        ws.cell(r, 1, label)
        ws.cell(r, 2, value)
        ws.cell(r, 1).font = font(10, True, NAVY)
        ws.cell(r, 1).fill = fill(LIGHT_GRAY)
        ws.cell(r, 2).fill = fill(PALE_ORANGE if r in (4, 5) else WHITE)
        ws.cell(r, 2).font = font(10, True, BLUE if r in (4, 5) else GREEN)
        ws.cell(r, 1).border = ws.cell(r, 2).border = grid_border
    ws["B4"].number_format = "0.00"
    ws["B4"].comment = Comment("ลองปรับ 0.10–0.80 แล้วสังเกตกราฟ Loss ในชีต 03_Model_Path", "OpenAI")
    ws["B5"].comment = Comment(f"ไฟล์นี้เตรียมพื้นที่ฝึกไว้ {TRAINING_ITERATIONS} รอบ", "OpenAI")
    ws["C7"] = "b เริ่มต้น"
    ws["D7"] = 0
    ws["E7"] = "w1 เริ่มต้น"
    ws["F7"] = 0
    ws["G7"] = "w2 เริ่มต้น"
    ws["H7"] = 0
    for c in range(3, 9):
        ws.cell(7, c).border = grid_border
        ws.cell(7, c).alignment = Alignment(horizontal="center")
        if c % 2:
            ws.cell(7, c).font = font(10, True, NAVY)
            ws.cell(7, c).fill = fill(LIGHT_GRAY)
        else:
            ws.cell(7, c).font = font(10, True, BLUE)
            ws.cell(7, c).fill = fill(PALE_ORANGE)

    dv_lr = DataValidation(type="decimal", operator="between", formula1="0.01", formula2="1", allow_blank=False)
    dv_lr.promptTitle = "Learning rate"
    dv_lr.prompt = "กรอกค่าระหว่าง 0.01 ถึง 1.00"
    dv_lr.error = "Learning rate ต้องอยู่ระหว่าง 0.01 ถึง 1.00"
    dv_lr.errorTitle = "ค่าที่กรอกไม่ถูกต้อง"
    dv_lr.showErrorMessage = True
    ws.add_data_validation(dv_lr)
    dv_lr.add(ws["B4"])

    if PRACTICE_MODE:
        note_box(
            ws,
            9,
            1,
            11,
            "วิธีทำ: ข้อมูลพื้นสีน้ำเงินเป็นโจทย์ที่กำหนดให้ • เติมสูตรหรือค่าที่คำนวณได้ในช่องสีเหลือง • ทำ Summary และ Update ก่อนขึ้นรอบถัดไป",
            PALE_ORANGE,
        )

    block_meta = []
    train_start = 6
    for iteration in range(1, TRAINING_ITERATIONS + 1):
        start = 11 + (iteration - 1) * 26
        param_row = start + 1
        head_row = start + 2
        first_data = start + 3
        last_data = first_data + 19
        summary_row = last_data + 1
        new_row = summary_row + 1
        block_meta.append((param_row, summary_row, new_row))

        section(
            ws,
            start,
            f"รอบที่ {iteration:02d}  •  พารามิเตอร์ปัจจุบัน → z → sigmoid p → error → gradient → log loss → พารามิเตอร์ใหม่",
            1,
            11,
            TEAL if iteration % 2 else NAVY,
        )
        ws.cell(param_row, 1, "Iteration")
        ws.cell(param_row, 2, iteration)
        ws.cell(param_row, 3, "b")
        ws.cell(param_row, 5, "w1")
        ws.cell(param_row, 7, "w2")
        if iteration == 1:
            ws.cell(param_row, 4, "=$D$7")
            ws.cell(param_row, 6, "=$F$7")
            ws.cell(param_row, 8, "=$H$7")
        else:
            prev_new = block_meta[-2][2]
            if PRACTICE_MODE:
                ws.cell(param_row, 4, f'=IF(D{prev_new}="","",D{prev_new})')
                ws.cell(param_row, 6, f'=IF(F{prev_new}="","",F{prev_new})')
                ws.cell(param_row, 8, f'=IF(H{prev_new}="","",H{prev_new})')
            else:
                ws.cell(param_row, 4, f"=D{prev_new}")
                ws.cell(param_row, 6, f"=F{prev_new}")
                ws.cell(param_row, 8, f"=H{prev_new}")
        ws.merge_cells(start_row=param_row, start_column=9, end_row=param_row, end_column=11)
        ws.cell(param_row, 9, "ใช้ค่าน้ำหนักชุดนี้กับข้อมูล Train ทุกแถว")
        for c in range(1, 12):
            cell = ws.cell(param_row, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = fill(PALE_BLUE if c in (4, 6, 8) else LIGHT_GRAY)
            cell.font = font(9, True if c in (1, 3, 5, 7) else False, NAVY)

        headers = ["ID", "x1", "x2", "y", "z", "p = sigmoid(z)", "error = p-y", "grad b = error", "grad w1 = error×x1", "grad w2 = error×x2", "Log Loss"]
        for c, value in enumerate(headers, 1):
            ws.cell(head_row, c, value)
        header_row(ws, head_row, 1, 11)

        for idx, r in enumerate(range(first_data, last_data + 1)):
            data_row = train_start + idx
            ws.cell(r, 1, f"='01_Data'!A{data_row}")
            ws.cell(r, 2, f"='01_Data'!F{data_row}")
            ws.cell(r, 3, f"='01_Data'!G{data_row}")
            ws.cell(r, 4, f"='01_Data'!D{data_row}")
            if not PRACTICE_MODE:
                ws.cell(r, 5, f"=$D${param_row}+$F${param_row}*B{r}+$H${param_row}*C{r}")
                ws.cell(r, 6, f"=1/(1+EXP(-E{r}))")
                ws.cell(r, 7, f"=F{r}-D{r}")
                ws.cell(r, 8, f"=G{r}")
                ws.cell(r, 9, f"=G{r}*B{r}")
                ws.cell(r, 10, f"=G{r}*C{r}")
                ws.cell(r, 11, f"=-(D{r}*LN(MIN(MAX(F{r},0.000001),0.999999))+(1-D{r})*LN(MIN(MAX(1-F{r},0.000001),0.999999)))")
            for c in range(1, 12):
                cell = ws.cell(r, c)
                cell.border = grid_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font(9, False, GREEN if c <= 4 else (BLUE if PRACTICE_MODE else BLACK))
                if c <= 4:
                    cell.fill = fill(SKY)
                elif PRACTICE_MODE:
                    cell.fill = fill(PALE_ORANGE)
                elif c in (7, 8, 9, 10):
                    cell.fill = fill(LIGHT_GRAY)
                else:
                    cell.fill = fill(WHITE)
            for c in range(2, 12):
                ws.cell(r, c).number_format = "0.0000"

        ws.cell(summary_row, 1, "สรุป Gradient และ Loss")
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)
        ws.cell(summary_row, 7, "ค่าเฉลี่ย")
        if not PRACTICE_MODE:
            ws.cell(summary_row, 8, f"=AVERAGE(H{first_data}:H{last_data})")
            ws.cell(summary_row, 9, f"=AVERAGE(I{first_data}:I{last_data})")
            ws.cell(summary_row, 10, f"=AVERAGE(J{first_data}:J{last_data})")
            ws.cell(summary_row, 11, f"=AVERAGE(K{first_data}:K{last_data})")
        for c in range(1, 12):
            cell = ws.cell(summary_row, c)
            cell.border = grid_border
            cell.fill = fill(PALE_ORANGE if PRACTICE_MODE and c >= 8 else PALE_PURPLE)
            cell.font = font(9, True, BLUE if PRACTICE_MODE and c >= 8 else PURPLE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c >= 8:
                cell.number_format = "0.0000"

        ws.cell(new_row, 1, "Update: พารามิเตอร์ใหม่ = ค่าเดิม - α × gradient")
        ws.merge_cells(start_row=new_row, start_column=1, end_row=new_row, end_column=2)
        ws.cell(new_row, 3, "b ใหม่")
        if not PRACTICE_MODE:
            ws.cell(new_row, 4, f"=D{param_row}-$B$4*H{summary_row}")
        ws.cell(new_row, 5, "w1 ใหม่")
        if not PRACTICE_MODE:
            ws.cell(new_row, 6, f"=F{param_row}-$B$4*I{summary_row}")
        ws.cell(new_row, 7, "w2 ใหม่")
        if not PRACTICE_MODE:
            ws.cell(new_row, 8, f"=H{param_row}-$B$4*J{summary_row}")
        ws.merge_cells(start_row=new_row, start_column=9, end_row=new_row, end_column=11)
        ws.cell(new_row, 9, "ค่าชุดนี้จะเป็นพารามิเตอร์ปัจจุบันของรอบถัดไป")
        for c in range(1, 12):
            cell = ws.cell(new_row, c)
            cell.border = grid_border
            cell.fill = fill(PALE_ORANGE if PRACTICE_MODE and c in (4, 6, 8) else (PALE_GREEN if c in (4, 6, 8) else LIGHT_GRAY))
            cell.font = font(9, True if c in (1, 3, 5, 7) else False, BLUE if PRACTICE_MODE and c in (4, 6, 8) else (GREEN if c in (4, 6, 8) else NAVY))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c in (4, 6, 8):
                cell.number_format = "0.0000"

    ws["E14"].comment = Comment("z คือคะแนนเชิงเส้นก่อนผ่าน sigmoid", "OpenAI")
    ws["F14"].comment = Comment("sigmoid แปลง z ให้เป็นความน่าจะเป็นระหว่าง 0 และ 1", "OpenAI")
    ws["G14"].comment = Comment("error = p - y เป็นทิศทางความคลาดเคลื่อนของแถวนั้น", "OpenAI")
    ws["K14"].comment = Comment("Log Loss ลงโทษการทำนายที่มั่นใจแต่ผิดมากเป็นพิเศษ", "OpenAI")
    ws.print_title_rows = "1:10"
    ws.print_area = f"A1:K{11 + (TRAINING_ITERATIONS - 1) * 26 + 24}"
    return block_meta


def build_model_path(wb, block_meta):
    ws = wb.create_sheet("03_Model_Path")
    configure_sheet(ws)
    title(
        ws,
        "03 • เส้นทางการเรียนรู้ของโมเดล",
        f"ติดตามค่าน้ำหนัก Gradient และ Average Log Loss ตลอด {TRAINING_ITERATIONS} รอบ",
        20,
    )
    ws.freeze_panes = "A6"
    headers = ["Iteration", "b ก่อน", "w1 ก่อน", "w2 ก่อน", "Avg Loss", "grad b", "grad w1", "grad w2", "b หลัง", "w1 หลัง", "w2 หลัง"]
    for c, value in enumerate(headers, 1):
        ws.cell(5, c, value)
    header_row(ws, 5, 1, 11)
    for col in "ABCDEFGHIJK":
        ws.column_dimensions[col].width = 13
    for i, (param_row, summary_row, new_row) in enumerate(block_meta, 6):
        ws.cell(i, 1, i - 5)
        refs = [
            f"='02_Training_Detail'!D{param_row}", f"='02_Training_Detail'!F{param_row}", f"='02_Training_Detail'!H{param_row}",
            f"='02_Training_Detail'!K{summary_row}", f"='02_Training_Detail'!H{summary_row}", f"='02_Training_Detail'!I{summary_row}", f"='02_Training_Detail'!J{summary_row}",
            f"='02_Training_Detail'!D{new_row}", f"='02_Training_Detail'!F{new_row}", f"='02_Training_Detail'!H{new_row}",
        ]
        for c, formula in enumerate(refs, 2):
            if PRACTICE_MODE:
                source_ref = formula[1:]
                ws.cell(i, c, f'=IF({source_ref}="","",{source_ref})')
            else:
                ws.cell(i, c, formula)
        for c in range(1, 12):
            ws.cell(i, c).border = grid_border
            ws.cell(i, c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(i, c).font = font(9, False, GREEN if c > 1 else NAVY)
            ws.cell(i, c).fill = fill(SKY if i % 2 == 0 else WHITE)
            if c > 1:
                ws.cell(i, c).number_format = "0.0000"
        ws.row_dimensions[i].height = 23

    data_end_row = 5 + TRAINING_ITERATIONS
    final_section_row = data_end_row + 3
    section(ws, final_section_row, "โมเดลสุดท้าย", 1, 5, ORANGE)
    finals = [
        (final_section_row + 1, "Final b", f"=I{data_end_row}"),
        (final_section_row + 2, "Final w1", f"=J{data_end_row}"),
        (final_section_row + 3, "Final w2", f"=K{data_end_row}"),
    ]
    for r, label, formula_value in finals:
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula_value)
        ws.cell(r, 1).font = font(10, True, NAVY)
        ws.cell(r, 1).fill = fill(LIGHT_GRAY)
        ws.cell(r, 2).font = font(10, True, BLACK)
        ws.cell(r, 2).fill = fill(PALE_GREEN)
        ws.cell(r, 1).border = ws.cell(r, 2).border = grid_border
        ws.cell(r, 2).number_format = "0.0000"
    equation_rows = [
        (final_section_row + 1, "สมการสุดท้าย: z = w1x1 + w2x2 + b", 11, True, PURPLE, PALE_PURPLE),
        (final_section_row + 2, "Probability: p = 1 / (1 + EXP(-z))", 11, True, PURPLE, PALE_PURPLE),
        (final_section_row + 3, "ถ้า p ≥ threshold → ทำนาย 1, ถ้า p < threshold → ทำนาย 0", 10, False, NAVY, SKY),
    ]
    for row, text_value, size, bold, color, background in equation_rows:
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=11)
        cell = ws.cell(row, 4, text_value)
        cell.font = font(size, bold, color)
        cell.fill = fill(background)
        cell.border = grid_border
        cell.alignment = Alignment(vertical="center")

    loss_chart = LineChart()
    loss_chart.title = "Average Log Loss ลดลงในแต่ละรอบ"
    loss_chart.y_axis.title = "Average Log Loss"
    loss_chart.x_axis.title = "Iteration"
    loss_chart.style = 13
    loss_chart.height = 7.5
    loss_chart.width = 15
    loss_chart.add_data(Reference(ws, min_col=5, min_row=5, max_row=data_end_row), titles_from_data=True)
    loss_chart.set_categories(Reference(ws, min_col=1, min_row=6, max_row=data_end_row))
    loss_chart.legend = None
    ws.add_chart(loss_chart, "M5")

    weight_chart = LineChart()
    weight_chart.title = "การเปลี่ยนแปลงของ b, w1, w2"
    weight_chart.y_axis.title = "Weight value"
    weight_chart.x_axis.title = "Iteration"
    weight_chart.style = 12
    weight_chart.height = 7.5
    weight_chart.width = 15
    weight_chart.add_data(Reference(ws, min_col=9, max_col=11, min_row=5, max_row=data_end_row), titles_from_data=True)
    weight_chart.set_categories(Reference(ws, min_col=1, min_row=6, max_row=data_end_row))
    ws.add_chart(weight_chart, "M21")
    ws.print_area = f"A1:U{max(36, final_section_row + 10)}"


def build_prediction(wb):
    ws = wb.create_sheet("04_Test_Prediction")
    configure_sheet(ws)
    title(
        ws,
        "04 • ทำนายข้อมูล Test",
        "แยก contribution ของ w1x1, w2x2 และ b เพื่อให้เห็นที่มาของ z และ probability",
        23,
    )
    ws.freeze_panes = "A8"
    ws["A4"] = "Threshold"
    ws["B4"] = 0.5
    ws["A4"].font = font(10, True, NAVY)
    ws["A4"].fill = fill(LIGHT_GRAY)
    ws["B4"].font = font(11, True, BLUE)
    ws["B4"].fill = fill(PALE_ORANGE)
    ws["B4"].number_format = "0.00"
    ws["A4"].border = ws["B4"].border = grid_border
    final_value_row = TRAINING_ITERATIONS + 9
    labels = [
        ("D4", "Final b", "E4", f"='03_Model_Path'!B{final_value_row}"),
        ("G4", "Final w1", "H4", f"='03_Model_Path'!B{final_value_row + 1}"),
        ("J4", "Final w2", "K4", f"='03_Model_Path'!B{final_value_row + 2}"),
    ]
    for label_cell, label, value_cell, formula_value in labels:
        ws[label_cell] = label
        if PRACTICE_MODE:
            source_ref = formula_value[1:]
            ws[value_cell] = f'=IF({source_ref}="","",{source_ref})'
        else:
            ws[value_cell] = formula_value
        ws[label_cell].font = font(10, True, NAVY)
        ws[label_cell].fill = fill(LIGHT_GRAY)
        ws[value_cell].font = font(10, True, GREEN)
        ws[value_cell].fill = fill(SKY)
        ws[label_cell].border = ws[value_cell].border = grid_border
        ws[value_cell].number_format = "0.0000"
    ws["B4"].comment = Comment("ลองเปลี่ยนเป็น 0.40 หรือ 0.60 แล้วดูผลที่ชีต 05_Evaluation", "OpenAI")
    dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=False)
    dv.promptTitle = "Threshold"
    dv.prompt = "กรอกค่าระหว่าง 0 และ 1"
    dv.error = "Threshold ต้องอยู่ระหว่าง 0 และ 1"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(ws["B4"])

    headers = [
        "ID", "Hours", "Attend %", "x1", "x2", "Actual y", "b", "w1×x1", "w2×x2", "z", "Probability p", "Predicted", "Outcome", "Correct", "Threshold",
    ]
    for c, value in enumerate(headers, 1):
        ws.cell(7, c, value)
    header_row(ws, 7, 1, 15)
    widths = [11, 10, 11, 11, 11, 11, 11, 12, 12, 11, 15, 13, 11, 11, 12]
    for c, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = width
    for idx, r in enumerate(range(8, 18)):
        data_row = 26 + idx
        ws.cell(r, 1, f"='01_Data'!A{data_row}")
        ws.cell(r, 2, f"='01_Data'!B{data_row}")
        ws.cell(r, 3, f"='01_Data'!C{data_row}")
        ws.cell(r, 4, f"='01_Data'!F{data_row}")
        ws.cell(r, 5, f"='01_Data'!G{data_row}")
        ws.cell(r, 6, f"='01_Data'!D{data_row}")
        if PRACTICE_MODE:
            ws.cell(r, 7, '=IF($E$4="","",$E$4)')
            ws.cell(r, 8, f'=IF($H$4="","",$H$4*D{r})')
            ws.cell(r, 9, f'=IF($K$4="","",$K$4*E{r})')
            ws.cell(r, 10, f'=IF(COUNT(G{r}:I{r})<3,"",SUM(G{r}:I{r}))')
            ws.cell(r, 11, f'=IF(J{r}="","",1/(1+EXP(-J{r})))')
            ws.cell(r, 12, f'=IF(K{r}="","",IF(K{r}>=$B$4,1,0))')
            ws.cell(r, 13, f'=IF(L{r}="","",IF(AND(F{r}=1,L{r}=1),"TP",IF(AND(F{r}=0,L{r}=1),"FP",IF(AND(F{r}=1,L{r}=0),"FN","TN"))))')
            ws.cell(r, 14, f'=IF(L{r}="","",--(F{r}=L{r}))')
        else:
            ws.cell(r, 7, "=$E$4")
            ws.cell(r, 8, f"=$H$4*D{r}")
            ws.cell(r, 9, f"=$K$4*E{r}")
            ws.cell(r, 10, f"=SUM(G{r}:I{r})")
            ws.cell(r, 11, f"=1/(1+EXP(-J{r}))")
            ws.cell(r, 12, f"=IF(K{r}>=$B$4,1,0)")
            ws.cell(r, 13, f'=IF(AND(F{r}=1,L{r}=1),"TP",IF(AND(F{r}=0,L{r}=1),"FP",IF(AND(F{r}=1,L{r}=0),"FN","TN")))')
            ws.cell(r, 14, f"=--(F{r}=L{r})")
        ws.cell(r, 15, "=$B$4")
        for c in range(1, 16):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = font(9, False, GREEN if c <= 6 else BLACK)
            cell.fill = fill(SKY if c <= 6 else WHITE)
        for c in range(4, 12):
            ws.cell(r, c).number_format = "0.0000"
        ws.cell(r, 15).number_format = "0.00"
        ws.row_dimensions[r].height = 26
    ws.auto_filter.ref = "A7:O17"
    ws.conditional_formatting.add("M8:M17", FormulaRule(formula=['$M8="TP"'], fill=fill(PALE_GREEN), font=font(9, True, GREEN)))
    ws.conditional_formatting.add("M8:M17", FormulaRule(formula=['$M8="TN"'], fill=fill(PALE_GREEN), font=font(9, True, GREEN)))
    ws.conditional_formatting.add("M8:M17", FormulaRule(formula=['$M8="FP"'], fill=fill(PALE_RED), font=font(9, True, RED)))
    ws.conditional_formatting.add("M8:M17", FormulaRule(formula=['$M8="FN"'], fill=fill(PALE_RED), font=font(9, True, RED)))
    ws.conditional_formatting.add("N8:N17", CellIsRule(operator="equal", formula=[1], fill=fill(PALE_GREEN)))
    ws.conditional_formatting.add("N8:N17", CellIsRule(operator="equal", formula=[0], fill=fill(PALE_RED)))

    chart = LineChart()
    chart.title = "Probability เทียบกับ Threshold"
    chart.y_axis.title = "Probability"
    chart.x_axis.title = "Student ID"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.style = 13
    chart.height = 8
    chart.width = 16
    chart.add_data(Reference(ws, min_col=11, max_col=11, min_row=7, max_row=17), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=15, max_col=15, min_row=7, max_row=17), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=8, max_row=17))
    ws.add_chart(chart, "Q7")
    note_box(ws, 20, 1, 15, "อ่าน Outcome: TP = ผ่านจริงและทำนายว่าผ่าน, TN = ไม่ผ่านจริงและทำนายว่าไม่ผ่าน, FP/FN คือกรณีที่โมเดลทำนายคลาดเคลื่อน", SKY)
    ws.print_area = "A1:W28"


def build_evaluation(wb):
    ws = wb.create_sheet("05_Evaluation")
    configure_sheet(ws)
    title(
        ws,
        "05 • Confusion Matrix และ Model Metrics",
        "คำนวณ Accuracy, Precision, Recall และ F1-score จากผลทำนายชุด Test",
        10,
    )
    for col, width in {"A": 4, "B": 22, "C": 18, "D": 18, "E": 16, "F": 4, "G": 20, "H": 25, "I": 25, "J": 4}.items():
        ws.column_dimensions[col].width = width
    ws["B4"] = "Threshold ปัจจุบัน"
    ws["C4"] = "='04_Test_Prediction'!B4"
    ws["B4"].font = font(10, True, NAVY)
    ws["B4"].fill = fill(LIGHT_GRAY)
    ws["C4"].font = font(11, True, GREEN)
    ws["C4"].fill = fill(PALE_ORANGE)
    ws["C4"].number_format = "0.00"
    ws["B4"].border = ws["C4"].border = grid_border

    section(ws, 6, "Confusion Matrix", 2, 5)
    ws["B7"] = "Actual / Predicted"
    ws["C7"] = "Predicted = 1"
    ws["D7"] = "Predicted = 0"
    ws["E7"] = "รวม Actual"
    ws["B8"] = "Actual = 1"
    ws["B9"] = "Actual = 0"
    ws["B10"] = "รวม Predicted"
    ws["C8"] = '=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,1,\'04_Test_Prediction\'!$L$8:$L$17,1)'
    ws["D8"] = '=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,1,\'04_Test_Prediction\'!$L$8:$L$17,0)'
    ws["C9"] = '=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,0,\'04_Test_Prediction\'!$L$8:$L$17,1)'
    ws["D9"] = '=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,0,\'04_Test_Prediction\'!$L$8:$L$17,0)'
    ws["E8"] = "=SUM(C8:D8)"
    ws["E9"] = "=SUM(C9:D9)"
    ws["C10"] = "=SUM(C8:C9)"
    ws["D10"] = "=SUM(D8:D9)"
    ws["E10"] = "=SUM(C8:D9)"
    for r in range(7, 11):
        for c in range(2, 6):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = font(10, True if r == 7 or c == 2 else False, NAVY)
            cell.fill = fill(LIGHT_GRAY if r == 7 or c == 2 else WHITE)
    ws["C8"].fill = fill(PALE_GREEN)
    ws["D9"].fill = fill(PALE_GREEN)
    ws["C9"].fill = fill(PALE_RED)
    ws["D8"].fill = fill(PALE_RED)
    ws["C8"].comment = Comment("True Positive (TP)", "OpenAI")
    ws["D8"].comment = Comment("False Negative (FN)", "OpenAI")
    ws["C9"].comment = Comment("False Positive (FP)", "OpenAI")
    ws["D9"].comment = Comment("True Negative (TN)", "OpenAI")
    ws["G7"] = "สัญลักษณ์"
    ws["H7"] = "ความหมาย"
    ws["G8"] = "TP"
    ws["H8"] = "ผ่านจริง และโมเดลทำนายว่าผ่าน"
    ws["G9"] = "FP"
    ws["H9"] = "ไม่ผ่านจริง แต่โมเดลทำนายว่าผ่าน"
    ws["G10"] = "FN"
    ws["H10"] = "ผ่านจริง แต่โมเดลทำนายว่าไม่ผ่าน"
    ws["G11"] = "TN"
    ws["H11"] = "ไม่ผ่านจริง และโมเดลทำนายว่าไม่ผ่าน"
    for r in range(7, 12):
        for c in range(7, 9):
            ws.cell(r, c).border = grid_border
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(r, c).font = font(9, True if r == 7 or c == 7 else False, NAVY)
            ws.cell(r, c).fill = fill(LIGHT_GRAY if r == 7 else WHITE)

    section(ws, 13, "Metrics", 2, 9, ORANGE)
    metrics_headers = ["Metric", "ค่า", "สูตรจาก Matrix", "คำถามที่ metric ตอบ"]
    for c, value in zip((2, 3, 4, 7), metrics_headers):
        ws.cell(14, c, value)
    ws.merge_cells("D14:F14")
    ws.merge_cells("G14:I14")
    header_row(ws, 14, 2, 9)
    metrics = [
        (15, "Accuracy", "=IFERROR((C8+D9)/SUM(C8:D9),0)", "(TP+TN)/(TP+TN+FP+FN)", "โดยรวมโมเดลทำนายถูกกี่เปอร์เซ็นต์?"),
        (16, "Precision", "=IFERROR(C8/(C8+C9),0)", "TP/(TP+FP)", "ในคนที่โมเดลบอกว่าผ่าน มีผ่านจริงกี่เปอร์เซ็นต์?"),
        (17, "Recall", "=IFERROR(C8/(C8+D8),0)", "TP/(TP+FN)", "ในคนที่ผ่านจริง โมเดลค้นพบได้กี่เปอร์เซ็นต์?"),
        (18, "F1-score", "=IFERROR(2*C16*C17/(C16+C17),0)", "2×Precision×Recall/(Precision+Recall)", "Precision และ Recall สมดุลกันดีเพียงใด?"),
        (19, "Error rate", "=1-C15", "1-Accuracy", "โดยรวมโมเดลทำนายผิดกี่เปอร์เซ็นต์?"),
    ]
    for r, label, formula_value, formula_text, meaning in metrics:
        ws.cell(r, 2, label)
        ws.cell(r, 3, formula_value)
        ws.cell(r, 4, formula_text)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(r, 7, meaning)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        for c in range(2, 10):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = font(10, True if c == 2 else False, NAVY if c == 2 else BLACK)
            cell.fill = fill(SKY if r % 2 else WHITE)
        ws.cell(r, 3).number_format = "0.0%"
        ws.cell(r, 3).font = font(12, True, TEAL)
        ws.row_dimensions[r].height = 34

    note_box(
        ws,
        21,
        2,
        9,
        "สรุปจะเปลี่ยนอัตโนมัติเมื่อแก้ threshold: "
        '="Accuracy "&TEXT(C15,"0.0%")&" | Precision "&TEXT(C16,"0.0%")&" | Recall "&TEXT(C17,"0.0%")&" | F1 "&TEXT(C18,"0.0%")',
        PALE_GREEN,
    )
    ws["B21"] = '="ที่ threshold "&TEXT(C4,"0.00")&" → Accuracy "&TEXT(C15,"0.0%")&" | Precision "&TEXT(C16,"0.0%")&" | Recall "&TEXT(C17,"0.0%")&" | F1 "&TEXT(C18,"0.0%")'
    ws["B21"].font = font(11, True, GREEN)
    ws["B21"].fill = fill(PALE_GREEN)
    ws["B21"].alignment = Alignment(vertical="center", wrap_text=True)
    ws["B21"].border = Border(left=medium_navy, right=thin_gray, top=thin_gray, bottom=thin_gray)
    ws.print_area = "A1:J24"


def build_key_concepts(wb):
    ws = wb.create_sheet("06_Key_Concepts")
    configure_sheet(ws)
    title(
        ws,
        "06 • แนวคิดสำคัญที่ต้องเข้าใจ",
        "คำอธิบายว่าแต่ละขั้นตอนทำอะไร ทำไปทำไม และควรตีความอย่างไร",
        10,
    )
    ws.freeze_panes = "A8"
    widths = {"A": 9, "B": 17, "C": 17, "D": 20, "E": 20, "F": 20, "G": 21, "H": 21, "I": 21, "J": 21}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    note_box(
        ws,
        4,
        1,
        10,
        "ภาพรวม: ข้อมูล Train ใช้เรียนรู้ค่าน้ำหนัก → โมเดลแปลงข้อมูลเป็น z และ probability → threshold เปลี่ยน probability เป็น class → ข้อมูล Test ใช้วัดผลด้วย Confusion Matrix และ Metrics",
        SKY,
    )
    section(ws, 7, "คำอธิบายแต่ละแนวคิด", 1, 10)
    header_labels = {1: "ขั้น", 2: "แนวคิด", 4: "ความหมาย", 7: "ทำไมจึงสำคัญ / ข้อควรจำ"}
    for col, value in header_labels.items():
        ws.cell(8, col, value)
    header_row(ws, 8, 1, 10)
    ws.merge_cells("B8:C8")
    ws.merge_cells("D8:F8")
    ws.merge_cells("G8:J8")

    concepts = [
        ("1", "Train / Test split", "Train ใช้ปรับ weights w1, w2 และ bias b ส่วน Test กันไว้เพื่อวัดผลหลังฝึก", "ห้ามใช้ Test ฝึกโมเดล เพราะจะทำให้คะแนนประเมินดูดีเกินจริงและไม่สะท้อนข้อมูลใหม่"),
        ("2", "Feature scaling", "ทำให้ฟีเจอร์ที่มีหน่วยต่างกันมีขนาดใกล้กัน เช่น Hours 1–9 กับ Attendance 55–95", "ช่วยให้ Gradient Descent เสถียรและไม่ให้ฟีเจอร์หนึ่งดูสำคัญเพียงเพราะตัวเลขมีหน่วยใหญ่กว่า"),
        ("3", "Center", "ค่าดิบที่ต้องการให้แปลงเป็น 0 เช่น Hours center = 5", "เมื่อ x = 0 ส่วนของฟีเจอร์นั้นจะไม่เพิ่มหรือลด z ทำให้ bias/intercept b ตีความง่ายขึ้น"),
        ("4", "Scale", "ระยะของค่าดิบที่นับเป็น 1 หน่วยหลัง scaling เช่น Hours scale = 2.5", "Scale ไม่ใช่ค่าสูงสุด แต่เป็นตัวกำหนดขนาดของหนึ่งก้าวบนแกนฟีเจอร์"),
        ("5", "Linear score z", "z = w1x1 + w2x2 + b เป็นคะแนนรวมก่อนแปลงเป็น probability", "z ยังไม่ใช่ความน่าจะเป็นและอาจน้อยกว่า 0 หรือมากกว่า 1 ได้"),
        ("6", "Sigmoid", "p = 1/(1+EXP(-z)) แปลง z เป็นค่าระหว่าง 0 ถึง 1", "p ใช้สื่อระดับความมั่นใจของโมเดล แต่ไม่ใช่ class จนกว่าจะเลือก threshold"),
        ("7", "Log Loss", "วัดระยะห่างระหว่าง probability กับคำตอบจริง y และลงโทษการมั่นใจผิดอย่างมาก", "ใช้ probability โดยตรง จึงละเอียดกว่าแค่ดูว่าทำนาย class ถูกหรือผิด"),
        ("8", "Gradient", "ค่าเฉลี่ยของ error และ error×x บอกทิศทางที่ควรปรับน้ำหนัก", "เครื่องหมายของ Gradient บอกทิศทาง ส่วนขนาดบอกความแรงของการปรับ"),
        ("9", "Learning rate α", "กำหนดขนาดก้าวในการอัปเดตพารามิเตอร์ใหม่ = ค่าเดิม - α×gradient", "เล็กเกินไปเรียนช้า ใหญ่เกินไปอาจแกว่งหรือทำให้ Loss เพิ่มขึ้น"),
        ("10", "Threshold", "กติกาเปลี่ยน probability เป็น class เช่น p ≥ 0.50 ให้ทำนาย 1", "Threshold ไม่ได้ถูกเรียนรู้ในตัวอย่างนี้ และการเปลี่ยน threshold ทำให้ FP/FN เปลี่ยน"),
        ("11", "Confusion Matrix", "นับ TP, FP, FN, TN เพื่อแยกประเภทการทำนายถูกและผิด", "บอกชนิดของความผิดพลาดได้ละเอียดกว่า Accuracy เพียงค่าเดียว"),
        ("12", "Accuracy", "สัดส่วนที่ทำนายถูกทั้งหมด: (TP+TN)/(TP+TN+FP+FN)", "เข้าใจง่ายแต่ต้องระวัง หาก class หนึ่งมีจำนวนมากกว่าอีก class มาก Accuracy อาจดูสูงทั้งที่โมเดลพลาด class สำคัญ"),
        ("13", "Precision", "ในคนที่โมเดลทำนายเป็น 1 มีเป็น 1 จริงกี่ส่วน: TP/(TP+FP)", "สำคัญเมื่อ False Positive มีต้นทุนสูง"),
        ("14", "Recall", "ในคนที่เป็น 1 จริง โมเดลค้นพบได้กี่ส่วน: TP/(TP+FN)", "สำคัญเมื่อการพลาดกรณี Positive หรือ False Negative มีต้นทุนสูง"),
        ("15", "F1-score", "ค่าเฉลี่ยแบบ harmonic ของ Precision และ Recall", "ใช้เมื่ออยากให้ Precision และ Recall สมดุล โดยเฉพาะเมื่อจำนวนแต่ละ class ไม่เท่ากัน"),
    ]
    for r, (step_no, concept, meaning, importance) in enumerate(concepts, 9):
        ws.cell(r, 1, step_no)
        ws.cell(r, 2, concept)
        ws.cell(r, 4, meaning)
        ws.cell(r, 7, importance)
        for col in range(1, 11):
            c = ws.cell(r, col)
            c.border = grid_border
            c.alignment = Alignment(horizontal="center" if col == 1 else "left", vertical="center", wrap_text=True)
            c.fill = fill(SKY if r % 2 else WHITE)
            c.font = font(9, col in (1, 2), TEAL if col in (1, 2) else NAVY)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)
        ws.row_dimensions[r].height = 52

    section(ws, 24, "สรุปที่มาของค่าปรับสเกลใน 01_Data", 1, 10, ORANGE)
    scale_headers = {1: "พารามิเตอร์", 2: "ค่าในไฟล์", 3: "ที่มา", 6: "ความหมายหลังแปลง"}
    for col, value in scale_headers.items():
        ws.cell(25, col, value)
    header_row(ws, 25, 1, 10)
    ws.merge_cells("C25:E25")
    ws.merge_cells("F25:J25")
    scale_rows = [
        (26, "Hours center", "='01_Data'!K6", '="ค่าเฉลี่ย Train = "&TEXT(AVERAGE(\'01_Data\'!B6:B25),"0.00")', "5 ชั่วโมง → x1 = 0"),
        (27, "Hours scale", "='01_Data'!K7", '="Standard deviation จริง = "&TEXT(STDEVP(\'01_Data\'!B6:B25),"0.00")&" แล้วปัดเพื่อคำนวณง่าย"', "2.5/5/7.5 ชั่วโมง → -1/0/+1"),
        (28, "Attendance center", "='01_Data'!K8", '="ค่าเฉลี่ย Train = "&TEXT(AVERAGE(\'01_Data\'!C6:C25),"0.00")&" แล้วปัดเป็น 75"', "75% → x2 = 0"),
        (29, "Attendance scale", "='01_Data'!K9", '="เลือก 15 จุดเปอร์เซ็นต์; standard deviation จริง = "&TEXT(STDEVP(\'01_Data\'!C6:C25),"0.00")', "60/75/90% → -1/0/+1"),
    ]
    for r, label, selected, source, meaning in scale_rows:
        ws.cell(r, 1, label)
        ws.cell(r, 2, selected)
        ws.cell(r, 3, source)
        ws.cell(r, 6, meaning)
        for col in range(1, 11):
            c = ws.cell(r, col)
            c.border = grid_border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = fill(PALE_ORANGE if r % 2 else WHITE)
            c.font = font(9, col == 1, GREEN if col in (2, 3) else NAVY)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
        ws.cell(r, 2).number_format = "0.00"
        ws.row_dimensions[r].height = 38
    note_box(ws, 31, 1, 10, "คำสำคัญ: Center-and-Scale ในไฟล์นี้ใช้ค่าที่ใกล้สถิติจริงแต่ปัดให้คำนวณง่าย หากต้องการ z-score มาตรฐาน ให้ใช้ Mean และ Standard Deviation จริงโดยไม่ปัดค่า", PALE_PURPLE)
    ws.print_area = "A1:J33"


def build_exercises(wb):
    ws = wb.create_sheet("07_Exercises")
    configure_sheet(ws)
    title(
        ws,
        "07 • แบบฝึกหัดและการทดลอง Threshold",
        "ใช้ probability เดิม แต่จำแนก class ที่ threshold ต่างกัน เพื่อเปรียบเทียบ trade-off",
        12,
    )
    for col, width in {"A": 14, "B": 10, "C": 10, "D": 10, "E": 10, "F": 15, "G": 15, "H": 15, "I": 15, "J": 4, "K": 24, "L": 28}.items():
        ws.column_dimensions[col].width = width

    section(ws, 4, "ตารางทดลองอัตโนมัติ", 1, 9)
    headers = ["Threshold", "TP", "FP", "FN", "TN", "Accuracy", "Precision", "Recall", "F1-score"]
    for c, value in enumerate(headers, 1):
        ws.cell(5, c, value)
    header_row(ws, 5, 1, 9)
    thresholds = [0.30, 0.40, 0.50, 0.60, 0.70]
    for r, value in enumerate(thresholds, 6):
        ws.cell(r, 1, value)
        ws.cell(r, 2, f'=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,1,\'04_Test_Prediction\'!$K$8:$K$17,">="&A{r})')
        ws.cell(r, 3, f'=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,0,\'04_Test_Prediction\'!$K$8:$K$17,">="&A{r})')
        ws.cell(r, 4, f'=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,1,\'04_Test_Prediction\'!$K$8:$K$17,"<"&A{r})')
        ws.cell(r, 5, f'=COUNTIFS(\'04_Test_Prediction\'!$F$8:$F$17,0,\'04_Test_Prediction\'!$K$8:$K$17,"<"&A{r})')
        ws.cell(r, 6, f"=IFERROR((B{r}+E{r})/SUM(B{r}:E{r}),0)")
        ws.cell(r, 7, f"=IFERROR(B{r}/(B{r}+C{r}),0)")
        ws.cell(r, 8, f"=IFERROR(B{r}/(B{r}+D{r}),0)")
        ws.cell(r, 9, f"=IFERROR(2*G{r}*H{r}/(G{r}+H{r}),0)")
        for c in range(1, 10):
            cell = ws.cell(r, c)
            cell.border = grid_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = font(10, False, BLUE if c == 1 else BLACK)
            cell.fill = fill(PALE_ORANGE if c == 1 else (SKY if r % 2 == 0 else WHITE))
        ws.cell(r, 1).number_format = "0.00"
        for c in range(6, 10):
            ws.cell(r, c).number_format = "0.0%"
        ws.row_dimensions[r].height = 25
    ws.conditional_formatting.add("I6:I10", CellIsRule(operator="equal", formula=["MAX($I$6:$I$10)"], fill=fill(PALE_GREEN)))

    section(ws, 12, "คำถามสำหรับอภิปราย", 1, 12, ORANGE)
    questions = [
        "1. Threshold ใดให้ F1-score สูงที่สุด? เหตุใดจึงไม่จำเป็นต้องเป็น 0.50 เสมอไป?",
        "2. เมื่อเพิ่ม threshold จำนวน FP และ FN เปลี่ยนอย่างไร? อธิบายจากตาราง",
        "3. ถ้าโจทย์คือคัดกรองนักศึกษาที่เสี่ยงตก และไม่อยากพลาดผู้ที่ต้องช่วยเหลือ ควรเน้น metric ใด?",
        "4. กลับไปปรับ Learning rate ใน 02_Training_Detail เป็น 0.10 และ 0.80 แล้วเปรียบเทียบกราฟ Loss",
        "5. เปลี่ยนค่า y ของข้อมูล Train หนึ่งแถว แล้วสังเกตว่าค่าน้ำหนักสุดท้ายและ metrics เปลี่ยนอย่างไร",
    ]
    for r, question in enumerate(questions, 13):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        ws.cell(r, 1, question)
        ws.cell(r, 1).font = font(10, True if r == 13 else False, NAVY)
        ws.cell(r, 1).fill = fill(SKY if r % 2 else WHITE)
        ws.cell(r, 1).border = grid_border
        ws.cell(r, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 34
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
        ws.cell(r, 11, "บันทึกคำตอบ:")
        ws.cell(r, 11).font = font(9, False, GRAY, True)
        ws.cell(r, 11).fill = fill(PALE_ORANGE)
        ws.cell(r, 11).border = grid_border

    note_box(ws, 20, 1, 12, "แนวคิดสำคัญ: metric ที่ดีที่สุดขึ้นอยู่กับต้นทุนของ FP และ FN ในบริบทจริง ไม่ใช่เลือก Accuracy สูงสุดเพียงอย่างเดียว", PALE_PURPLE)
    ws.print_area = "A1:L22"


def build_workbook(output=OUT, iterations=15, practice=False):
    global TRAINING_ITERATIONS, PRACTICE_MODE
    TRAINING_ITERATIONS = iterations
    PRACTICE_MODE = practice

    if TRAINING_ITERATIONS < 1:
        raise ValueError("iterations must be at least 1")

    wb = Workbook()
    wb.properties.title = "Logistic Regression Practice - 20 Steps" if PRACTICE_MODE else "Logistic Regression Step by Step for Students"
    wb.properties.subject = "Excel-based Logistic Regression teaching workbook"
    wb.properties.creator = "OpenAI"
    wb.properties.description = (
        "Thai pre-exam practice workbook with blank student calculation cells"
        if PRACTICE_MODE
        else "Thai classroom workbook with formula-driven gradient descent, prediction, confusion matrix, and metrics"
    )
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb._named_styles["Normal"].font = font()

    build_guide(wb)
    build_data(wb)
    block_meta = build_training(wb)
    build_model_path(wb, block_meta)
    build_prediction(wb)
    build_evaluation(wb)
    build_key_concepts(wb)
    build_exercises(wb)
    wb.active = 0
    wb.save(output)
    print(output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build the Thai Logistic Regression teaching workbook")
    parser.add_argument("--iterations", type=int, default=15, help="number of gradient-descent training rounds")
    parser.add_argument("--practice", action="store_true", help="leave student-calculated training cells blank")
    parser.add_argument("--output", type=Path, help="output .xlsx path")
    args = parser.parse_args()
    selected_output = args.output
    if selected_output is None:
        selected_output = (
            Path(__file__).with_name(f"Logistic_Regression_Practice_{args.iterations}_Steps_TH.xlsx")
            if args.practice
            else OUT
        )
    build_workbook(selected_output, iterations=args.iterations, practice=args.practice)
